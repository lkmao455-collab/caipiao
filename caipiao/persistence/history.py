"""历史记录管理."""

from __future__ import annotations

import csv
import json
import logging
from collections.abc import Iterable
from datetime import datetime, timedelta, timezone
from pathlib import Path

from ..core.ticket import Ticket

logger = logging.getLogger(__name__)


class HistoryManager:
    """管理生成历史记录的增删改查与导入导出."""

    def __init__(self, storage_path: Path | str, max_entries: int = 1000) -> None:
        self.storage_path = Path(storage_path)
        try:
            self.max_entries = max(1, int(max_entries))
        except (ValueError, TypeError):
            self.max_entries = 1000
        self._tickets: list[Ticket] = []
        self._load()

    def _load(self) -> None:
        """从文件加载历史."""
        if not self.storage_path.exists():
            return
        try:
            with self.storage_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, list):
                raise TypeError("JSON root must be a list")
            self._tickets = [Ticket.from_dict(item) for item in data]
        except (OSError, json.JSONDecodeError, ValueError) as exc:
            logger.error("加载历史记录失败: %s", exc)
            self._tickets = []
        except TypeError as exc:
            logger.error("加载历史记录失败（类型错误）: %s", exc)
            self._tickets = []

    def save(self) -> None:
        """保存历史到文件."""
        self.storage_path.parent.mkdir(parents=True, exist_ok=True)
        data = [ticket.to_dict() for ticket in self._tickets[-self.max_entries :]]
        with self.storage_path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def add(self, ticket: Ticket, *, skip_duplicates: bool = False) -> bool:
        """添加单条记录。

        Args:
            ticket: 要添加的投注单。
            skip_duplicates: 为 True 时跳过与现有记录完全相同的票。

        Returns:
            是否实际添加了新记录。
        """
        if skip_duplicates and ticket in self._tickets:
            return False
        self._tickets.append(ticket)
        self._trim()
        self.save()
        return True

    def add_many(
        self, tickets: Iterable[Ticket], *, skip_duplicates: bool = False
    ) -> int:
        """批量添加记录。

        Args:
            tickets: 要添加的投注单列表。
            skip_duplicates: 为 True 时跳过已存在的记录。

        Returns:
            实际新增的记录数量。
        """
        added = 0
        for ticket in tickets:
            if skip_duplicates and ticket in self._tickets:
                continue
            self._tickets.append(ticket)
            added += 1
        if added > 0:
            self._trim()
            self.save()
        return added

    def _trim(self) -> None:
        """限制记录数量."""
        if len(self._tickets) > self.max_entries:
            self._tickets = self._tickets[-self.max_entries :]

    def get_all(self) -> list[Ticket]:
        """获取全部记录."""
        return self._tickets[:]

    def get_recent(self, days: int = 30) -> list[Ticket]:
        """获取最近 N 天的记录."""
        now = datetime.now(timezone.utc).astimezone()
        cutoff = now - timedelta(days=days)
        result = []
        for t in self._tickets:
            if not t.generated_at:
                continue
            gt = t.generated_at
            # 统一为带时区的 datetime 进行比较
            if gt.tzinfo is None:
                # 无时区 -> 视为本地时间
                gt = gt.replace(tzinfo=timezone.utc).astimezone()
            if gt >= cutoff:
                result.append(t)
        return result

    def clear(self) -> None:
        """清空历史."""
        self._tickets.clear()
        self.save()

    def delete(self, ticket: Ticket) -> bool:
        """删除指定记录."""
        try:
            self._tickets.remove(ticket)
            self.save()
            return True
        except ValueError:
            return False

    def export_csv(self, path: Path | str) -> None:
        """导出为 CSV."""
        path = Path(path)
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            # 动态列头：按第一个有号票的 render_groups 决定（全部票同彩种）
            sample = self._tickets[0] if self._tickets else None
            group_names = [rg.name for rg in sample.render_groups()] if sample else ["红球", "蓝球"]
            if not group_names:
                group_names = ["号码"]
            writer.writerow(["时间", "策略"] + group_names + ["紧凑格式", "依据"])
            for t in self._tickets:
                row = [
                    t.generated_at.strftime("%Y-%m-%d %H:%M:%S") if t.generated_at else "-",
                    t.strategy_name,
                ]
                groups = list(t.render_groups())
                if not groups:
                    row.append("")
                else:
                    for rg in groups:
                        row.append(" ".join(f"{n:0{rg.pad}d}" for n in rg.numbers))
                row.extend([t.format_compact(), t.basis])
                writer.writerow(row)

    def export_txt(self, path: Path | str) -> None:
        """导出为纯文本."""
        path = Path(path)
        with path.open("w", encoding="utf-8") as f:
            for t in self._tickets:
                f.write(t.format_compact() + "\n")

    def export_excel(self, path: Path | str) -> None:
        """导出为 Excel (.xlsx)。"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        path = Path(path)
        wb = Workbook()
        ws = wb.active
        ws.title = "历史记录"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 动态列头
        sample = self._tickets[0] if self._tickets else None
        group_names = [rg.name for rg in sample.render_groups()] if sample else ["号码"]
        headers = ["序号", "时间", "策略"] + group_names + ["紧凑格式", "依据"]

        # 写表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写数据
        for row_idx, t in enumerate(self._tickets, 2):
            row_data = [
                row_idx - 1,
                t.generated_at.strftime("%Y-%m-%d %H:%M:%S") if t.generated_at else "-",
                t.strategy_name,
            ]
            groups = list(t.render_groups())
            if groups:
                for rg in groups:
                    row_data.append(" ".join(f"{n:0{rg.pad}d}" for n in rg.numbers))
            else:
                row_data.append("")
            row_data.extend([t.format_compact(), t.basis])

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")

        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, len(self._tickets) + 2)
            )
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = (
                min(max_length + 4, 40)
            )

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(path)

    def import_from_json(self, path: Path | str) -> int:
        """从 JSON 导入历史记录."""
        path = Path(path)
        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, list):
                raise TypeError("JSON root must be a list")
            imported = [Ticket.from_dict(item) for item in data]
        except (OSError, json.JSONDecodeError, ValueError) as exc:
            logger.error("导入历史记录失败: %s", exc)
            return 0
        except TypeError as exc:
            logger.error("导入历史记录失败（类型错误）: %s", exc)
            return 0
        self.add_many(imported)
        return len(imported)
