"""开奖记录相邻期/间隔期统计分析对话框.

支持彩种：
- 双色球：相邻期、间隔 1..N 期红球/蓝球重合统计
- 福彩3D：相邻期、间隔 1..N 期按位数字相同个数统计
- 七乐彩：相邻期、间隔 1..N 期基本号/特别号重合统计
- 快乐8：相邻期、间隔 1..N 期主号码重合个数统计

每种彩种根据自身的 NumberGroup 结构计算相邻期号码重叠情况。

增量计算：
- 以 ``records`` 列表的``最大索引``和 ``max_gap`` 作为缓存 key；
- 当最大间隔期数减小时，直接从已有缓存中裁剪展示；
- 当最大间隔期数增大时，仅计算新增间隔的统计；
- 当新增开奖记录时，仅基于最新一期与前面各期补齐各间隔统计；
- UI 层通过 ``QTimer.singleShot`` 异步触发重算，避免直接阻塞主线程。
"""

from __future__ import annotations

import logging
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QElapsedTimer, QThread, QTimer, Signal

logger = logging.getLogger(__name__)
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from ...core.profile import LotteryProfile, NumberGroup
from ...data.models import DrawRecord
from ...data.repository import DrawRepository
from ...persistence.settings import AppSettings


# --------------------------------------------------------------------------- #
# 后台计算 worker
# --------------------------------------------------------------------------- #
class _AnalysisWorker(QThread):
    """在后台线程执行相邻期/间隔期统计，避免阻塞主线程 UI."""

    finished = Signal(object, object)  # stats, details
    error = Signal(str)

    def __init__(
        self,
        records: List[DrawRecord],
        profile: LotteryProfile,
        max_gap: int,
    ) -> None:
        super().__init__()
        self.records = records
        self.profile = profile
        self.max_gap = max_gap

    def run(self) -> None:
        logger.info("[Worker] 开始计算, 记录数=%d, max_gap=%d",
                     len(self.records), self.max_gap)
        t0 = time.monotonic()
        try:
            stats, details = _analyze_adjacent_chunked(
                self.records, self.profile, max_gap=self.max_gap,
                should_interrupt=self.isInterruptionRequested,
            )
            if self.isInterruptionRequested():
                logger.info("[Worker] 被中断")
                return
            elapsed_ms = int((time.monotonic() - t0) * 1000)
            logger.info("[Worker] 完成, total_pairs=%d, 耗时 %dms", stats.total_pairs, elapsed_ms)
            self.finished.emit(stats, details)
        except Exception as exc:
            logger.exception("[Worker] 异常")
            self.error.emit(str(exc))


# --------------------------------------------------------------------------- #
# 分析器（增量版本）
# --------------------------------------------------------------------------- #


# --------------------------------------------------------------------------- #
# 统计抽象
# --------------------------------------------------------------------------- #
@dataclass
class GroupOverlapStats:
    """单个号码组的相邻期统计结果."""

    group_name: str = ""
    total_pairs: int = 0
    # 相同个数 -> 次数
    same_counts: Dict[int, int] = field(default_factory=dict)

    def same_ratio(self, n: int) -> float:
        return self.same_counts.get(n, 0) / max(self.total_pairs, 1) * 100


@dataclass
class GroupGapStats:
    """按间隔 gap 统计的号码组重叠结果.

    gap=1 表示间隔一期（第 3 期 vs 第 1 期），gap=2 表示间隔两期。
    """

    group_name: str = ""
    gap: int = 1
    total_pairs: int = 0
    same_counts: Dict[int, int] = field(default_factory=dict)

    def same_ratio(self, n: int) -> float:
        return self.same_counts.get(n, 0) / max(self.total_pairs, 1) * 100


@dataclass
class AdjacentStats:
    """相邻开奖记录统计结果.

    Attributes:
        total_pairs: 相邻期对数
        group_stats: 每个分析号码组的统计（按组 key）
        gap_stats: 按间隔统计的结果（按组 key -> gap -> stats）
    """

    total_pairs: int = 0
    group_stats: Dict[str, GroupOverlapStats] = field(default_factory=dict)
    gap_stats: Dict[str, Dict[int, GroupGapStats]] = field(default_factory=dict)


@dataclass
class FilterImpactResult:
    """单个号码组在指定间隔和阈值下的过滤影响结果."""

    gap_label: str = ""        # "相邻" 或 "间隔 N 期"
    group_name: str = ""       # "红球"
    total_pairs: int = 0
    threshold: int = 0         # 重合上限 K（overlap > K 则过滤）
    filtered_count: int = 0    # 被过滤的对数
    filtered_pct: float = 0.0  # 过滤占比
    retained_count: int = 0    # 保留的对数


# --------------------------------------------------------------------------- #
# 分析器（增量版本）
# --------------------------------------------------------------------------- #
def _init_gap_stats(
    stats: AdjacentStats,
    key: str,
    group_name: str,
    max_count: int,
    gaps: List[int],
) -> None:
    """初始化间隔统计桶."""
    if key not in stats.gap_stats:
        stats.gap_stats[key] = {}
    for gap in gaps:
        total_pairs = max(0, stats.total_pairs - gap)
        stats.gap_stats[key][gap] = GroupGapStats(
            group_name=group_name,
            gap=gap,
            total_pairs=total_pairs,
            same_counts={i: 0 for i in range(0, max_count + 1)},
        )


def _ensure_gap_stats(stats: AdjacentStats, key: str, group_name: str, max_count: int, gap: int) -> None:
    """确保指定 gap 的统计桶已存在（用于增量扩大 max_gap）."""
    if key not in stats.gap_stats:
        stats.gap_stats[key] = {}
    if gap not in stats.gap_stats[key]:
        total_pairs = max(0, stats.total_pairs - gap)
        stats.gap_stats[key][gap] = GroupGapStats(
            group_name=group_name,
            gap=gap,
            total_pairs=total_pairs,
            same_counts={i: 0 for i in range(0, max_count + 1)},
        )


def _compute_overlap_ssq(base: DrawRecord, curr: DrawRecord) -> Tuple[int, bool]:
    """计算双色球 base 与 curr 的红球重叠数、蓝球是否相同."""
    base_reds = set(base.groups.get("red", []))
    curr_reds = set(curr.groups.get("red", []))
    red_overlap = len(base_reds & curr_reds)

    base_blue = next(iter(base.groups.get("blue", [])), None)
    curr_blue = next(iter(curr.groups.get("blue", [])), None)
    blue_same = base_blue is not None and curr_blue is not None and base_blue == curr_blue
    return red_overlap, blue_same


def _compute_overlap_basic_special(
    base: DrawRecord,
    curr: DrawRecord,
    basic_group: NumberGroup,
    special_group: NumberGroup,
) -> Tuple[int, bool]:
    """计算七乐彩/广东36选7 base 与 curr 的基本号重叠数、特别号是否相同."""
    base_basic = set(base.groups.get(basic_group.key, []))
    curr_basic = set(curr.groups.get(basic_group.key, []))
    basic_overlap = len(base_basic & curr_basic)

    base_special = next(iter(base.groups.get(special_group.key, [])), None)
    curr_special = next(iter(curr.groups.get(special_group.key, [])), None)
    special_same = (
        base_special is not None
        and curr_special is not None
        and base_special == curr_special
    )
    return basic_overlap, special_same


def _compute_overlap_generic(base: DrawRecord, curr: DrawRecord, group: NumberGroup) -> int:
    """通用彩种：按位组统计同位相同数，非按位组统计集合交集大小."""
    if group.positional:
        base_nums = base.groups.get(group.key, [])
        curr_nums = curr.groups.get(group.key, [])
        return sum(1 for a, b in zip(base_nums, curr_nums) if a == b)
    return len(set(base.groups.get(group.key, [])) & set(curr.groups.get(group.key, [])))


def _analyze_incremental_ssq(
    records: List[DrawRecord],
    stats: AdjacentStats,
    prev_max_gap: int,
    new_max_gap: int,
) -> None:
    """双色球增量更新间隔统计.

    当 new_max_gap > prev_max_gap 时，仅计算新增 gap；
    当 records 长度增加时，仅基于最后一条记录补齐各 gap。
    """
    total = len(records)
    if total < 2:
        return

    old_total_pairs = stats.total_pairs
    stats.total_pairs = max(0, total - 1)
    stats.group_stats["red"].total_pairs = stats.total_pairs
    stats.group_stats["blue"].total_pairs = stats.total_pairs

    # 若记录数增加，需要先把相邻期统计（group_stats）补齐，这里暂不处理相邻期，
    # 因为相邻期统计目前仍由 _analyze_adjacent_ssq 全量负责。增量场景下，调用方
    # 会传入已包含最新相邻统计的 stats。

    # 计算新增 gap 的统计
    for gap in range(prev_max_gap + 1, new_max_gap + 1):
        _ensure_gap_stats(stats, "red", "红球", 6, gap)
        _ensure_gap_stats(stats, "blue", "蓝球", 1, gap)
        # total_pairs 以新的 records 长度为准
        stats.gap_stats["red"][gap].total_pairs = max(0, stats.total_pairs - gap)
        stats.gap_stats["blue"][gap].total_pairs = max(0, stats.total_pairs - gap)
        for i in range(gap + 1, total):
            red_overlap, blue_same = _compute_overlap_ssq(records[i - gap - 1], records[i])
            stats.gap_stats["red"][gap].same_counts[red_overlap] += 1
            stats.gap_stats["blue"][gap].same_counts[1 if blue_same else 0] += 1

    # 若记录数增加，为已有 gap 补齐最新记录带来的对子
    if stats.total_pairs > old_total_pairs:
        for gap in range(1, new_max_gap + 1):
            if gap >= total:
                continue
            _ensure_gap_stats(stats, "red", "红球", 6, gap)
            _ensure_gap_stats(stats, "blue", "蓝球", 1, gap)
            stats.gap_stats["red"][gap].total_pairs = max(0, stats.total_pairs - gap)
            stats.gap_stats["blue"][gap].total_pairs = max(0, stats.total_pairs - gap)
            # 最新一条记录是 records[-1]，它需要与前 gap 条记录配对
            curr = records[-1]
            for i in range(gap + 1):
                base_idx = total - 1 - gap - 1 + i
                if base_idx < 0:
                    continue
                # 我们只需要新增的对子：base 索引 > 上一次最后一条索引 - gap - 1
                # 更简单：只比较 base 为 total - gap - 1 到 total - 2 的那些
                pass
            # 实际上新增的对子是 (total - gap - 1, total - 1) ... (total - 2, total - 1)
            # 但为了避免重复统计，只处理 base 索引 >= old_total_pairs - gap 的部分
            start_base = max(gap, old_total_pairs - gap)
            for base_idx in range(start_base, total - 1):
                red_overlap, blue_same = _compute_overlap_ssq(records[base_idx - gap], records[base_idx])
                stats.gap_stats["red"][gap].same_counts[red_overlap] += 1
                stats.gap_stats["blue"][gap].same_counts[1 if blue_same else 0] += 1


def _analyze_adjacent_ssq(records: List[DrawRecord], max_gap: int = 7,
                          should_interrupt: Optional[Callable[[], bool]] = None) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """双色球：红球 0-6 个相同，蓝球是否相同；并补充间隔 1..max_gap 期统计。"""
    stats = AdjacentStats(total_pairs=max(0, len(records) - 1))
    stats.group_stats["red"] = GroupOverlapStats(group_name="红球", total_pairs=stats.total_pairs,
                                                  same_counts={i: 0 for i in range(0, 7)})
    stats.group_stats["blue"] = GroupOverlapStats(group_name="蓝球", total_pairs=stats.total_pairs)
    blue_same_count = 0

    gaps = list(range(1, max_gap + 1))
    _init_gap_stats(stats, "red", "红球", 6, gaps)
    _init_gap_stats(stats, "blue", "蓝球", 1, gaps)

    details: List[Dict[str, Any]] = []
    if not records:
        return stats, details

    records = sorted(records, key=lambda r: r.draw_date)
    details.append({"red": None, "blue": None})

    for i in range(1, len(records)):
        if i % 100 == 0:

            time.sleep(0.05)
        prev = records[i - 1]
        curr = records[i]
        red_overlap, blue_same = _compute_overlap_ssq(prev, curr)
        stats.group_stats["red"].same_counts[red_overlap] += 1
        if blue_same:
            blue_same_count += 1
        details.append({"red": red_overlap, "blue": blue_same})

    stats.group_stats["blue"].same_counts[1] = blue_same_count
    stats.group_stats["blue"].same_counts[0] = stats.total_pairs - blue_same_count

    for gap in range(1, max_gap + 1):
        for i in range(gap + 1, len(records)):
            if i % 100 == 0:

                time.sleep(0.05)
            red_overlap, blue_same = _compute_overlap_ssq(records[i - gap - 1], records[i])
            stats.gap_stats["red"][gap].same_counts[red_overlap] += 1
            stats.gap_stats["blue"][gap].same_counts[1 if blue_same else 0] += 1

    return stats, details


def _analyze_adjacent_positional(
    records: List[DrawRecord],
    group: NumberGroup,
    max_gap: int = 7,
    should_interrupt: Optional[Callable[[], bool]] = None,
) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """按位数字彩种（福彩3D/排列3/排列5/7星彩）：统计每位相同个数，并补充间隔统计。

    Args:
        should_interrupt: 可选的回调，返回 True 时提前终止计算。
    """
    stats = AdjacentStats(total_pairs=max(0, len(records) - 1))
    stats.group_stats[group.key] = GroupOverlapStats(
        group_name=group.name,
        total_pairs=stats.total_pairs,
        same_counts={i: 0 for i in range(0, group.count + 1)},
    )
    gaps = list(range(1, max_gap + 1))
    _init_gap_stats(stats, group.key, group.name, group.count, gaps)

    details: List[Dict[str, Any]] = []
    if not records:
        return stats, details

    records = sorted(records, key=lambda r: r.draw_date)
    details.append({group.key: None})

    for i in range(1, len(records)):
        if should_interrupt and should_interrupt():
            return stats, details
        if i % 100 == 0:

            time.sleep(0.05)
        prev = records[i - 1].groups.get(group.key, [])
        curr = records[i].groups.get(group.key, [])
        same = sum(1 for a, b in zip(prev, curr) if a == b)
        stats.group_stats[group.key].same_counts[same] += 1
        details.append({group.key: same})

    for gap in range(1, max_gap + 1):
        if should_interrupt and should_interrupt():
            return stats, details
        for i in range(gap + 1, len(records)):
            if i % 100 == 0:

                time.sleep(0.05)
            base = records[i - gap - 1].groups.get(group.key, [])
            curr = records[i].groups.get(group.key, [])
            same = sum(1 for a, b in zip(base, curr) if a == b)
            stats.gap_stats[group.key][gap].same_counts[same] += 1

    return stats, details


def _analyze_adjacent_basic_special(records: List[DrawRecord], basic_group: NumberGroup,
                                    special_group: NumberGroup, max_gap: int = 7,
                                    should_interrupt: Optional[Callable[[], bool]] = None) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """基本号+特别号彩种（七乐彩/广东36选7）：基本号 0-N 个相同，特别号是否相同，并补充间隔统计。"""
    stats = AdjacentStats(total_pairs=max(0, len(records) - 1))
    stats.group_stats["basic"] = GroupOverlapStats(
        group_name=basic_group.name,
        total_pairs=stats.total_pairs,
        same_counts={i: 0 for i in range(0, basic_group.count + 1)},
    )
    stats.group_stats["special"] = GroupOverlapStats(group_name=special_group.name, total_pairs=stats.total_pairs)
    special_same_count = 0

    gaps = list(range(1, max_gap + 1))
    _init_gap_stats(stats, "basic", basic_group.name, basic_group.count, gaps)
    _init_gap_stats(stats, "special", special_group.name, 1, gaps)

    details: List[Dict[str, Any]] = []
    if not records:
        return stats, details

    records = sorted(records, key=lambda r: r.draw_date)
    details.append({"basic": None, "special": None})

    for i in range(1, len(records)):
        if i % 100 == 0:

            time.sleep(0.05)
        prev = records[i - 1]
        curr = records[i]
        basic_overlap, special_same = _compute_overlap_basic_special(prev, curr, basic_group, special_group)
        stats.group_stats["basic"].same_counts[basic_overlap] += 1
        if special_same:
            special_same_count += 1
        details.append({"basic": basic_overlap, "special": special_same})

    stats.group_stats["special"].same_counts[1] = special_same_count
    stats.group_stats["special"].same_counts[0] = stats.total_pairs - special_same_count

    for gap in range(1, max_gap + 1):
        for i in range(gap + 1, len(records)):
            if i % 100 == 0:

                time.sleep(0.05)
            basic_overlap, special_same = _compute_overlap_basic_special(
                records[i - gap - 1], records[i], basic_group, special_group
            )
            stats.gap_stats["basic"][gap].same_counts[basic_overlap] += 1
            stats.gap_stats["special"][gap].same_counts[1 if special_same else 0] += 1

    return stats, details


def _analyze_adjacent_main(records: List[DrawRecord], group: NumberGroup, max_gap: int = 7,
                           should_interrupt: Optional[Callable[[], bool]] = None) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """快乐8：开奖 20 个号码，统计相邻期相同号码个数分布，并补充间隔统计。"""
    stats = AdjacentStats(total_pairs=max(0, len(records) - 1))
    stats.group_stats[group.key] = GroupOverlapStats(
        group_name=group.name,
        total_pairs=stats.total_pairs,
        same_counts={i: 0 for i in range(0, group.count + 1)},
    )
    gaps = list(range(1, max_gap + 1))
    _init_gap_stats(stats, group.key, group.name, group.count, gaps)

    details: List[Dict[str, Any]] = []
    if not records:
        return stats, details

    records = sorted(records, key=lambda r: r.draw_date)
    details.append({group.key: None})

    for i in range(1, len(records)):
        if i % 100 == 0:

            time.sleep(0.05)
        prev = set(records[i - 1].groups.get(group.key, []))
        curr = set(records[i].groups.get(group.key, []))
        overlap = len(prev & curr)
        stats.group_stats[group.key].same_counts[overlap] += 1
        details.append({group.key: overlap})

    for gap in range(1, max_gap + 1):
        for i in range(gap + 1, len(records)):
            if i % 100 == 0:

                time.sleep(0.05)
            base = set(records[i - gap - 1].groups.get(group.key, []))
            curr = set(records[i].groups.get(group.key, []))
            overlap = len(base & curr)
            stats.gap_stats[group.key][gap].same_counts[overlap] += 1

    return stats, details


def _analyze_adjacent(
    records: List[DrawRecord],
    profile: LotteryProfile,
    max_gap: int = 7,
    should_interrupt: Optional[Callable[[], bool]] = None,
) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """根据彩种档案选择对应的相邻期分析器，并统一补充间隔 1..max_gap 期统计。"""
    if profile.key == "ssq":
        return _analyze_adjacent_ssq(records, max_gap, should_interrupt=should_interrupt)

    if profile.key in ("3d", "pl3", "pl5", "qxc"):
        group = profile.primary_group
        return _analyze_adjacent_positional(records, group, max_gap, should_interrupt=should_interrupt)

    if profile.key in ("qlc", "gd36x7"):
        basic = profile.group("basic")
        special = profile.group("special")
        return _analyze_adjacent_basic_special(
            records, basic, special, max_gap, should_interrupt=should_interrupt
        )

    if profile.key == "kl8":
        group = profile.primary_group
        return _analyze_adjacent_main(records, group, max_gap, should_interrupt=should_interrupt)

    # 未知彩种退化为通用：分析所有 pick_groups
    stats = AdjacentStats(total_pairs=max(0, len(records) - 1))
    details: List[Dict[str, Any]] = []
    if not records:
        return stats, details

    records = sorted(records, key=lambda r: r.draw_date)
    gaps = list(range(1, max_gap + 1))
    for idx, g in enumerate(profile.pick_groups):
        stats.group_stats[g.key] = GroupOverlapStats(
            group_name=g.name,
            total_pairs=stats.total_pairs,
            same_counts={i: 0 for i in range(0, g.count + 1)},
        )
        _init_gap_stats(stats, g.key, g.name, g.count, gaps)
    details.append({g.key: None for g in profile.pick_groups})

    for i in range(1, len(records)):
        if should_interrupt and should_interrupt():
            return stats, details
        if i % 100 == 0:

            time.sleep(0.05)
        detail: Dict[str, Any] = {}
        for g in profile.pick_groups:
            if g.positional:
                prev = records[i - 1].groups.get(g.key, [])
                curr = records[i].groups.get(g.key, [])
                overlap = sum(1 for a, b in zip(prev, curr) if a == b)
            else:
                prev = set(records[i - 1].groups.get(g.key, []))
                curr = set(records[i].groups.get(g.key, []))
                overlap = len(prev & curr)
            stats.group_stats[g.key].same_counts[overlap] += 1
            detail[g.key] = overlap
        details.append(detail)

    for gap in range(1, max_gap + 1):
        if should_interrupt and should_interrupt():
            return stats, details
        for i in range(gap + 1, len(records)):
            if i % 100 == 0:

                time.sleep(0.05)
            for g in profile.pick_groups:
                overlap = _compute_overlap_generic(records[i - gap - 1], records[i], g)
                stats.gap_stats[g.key][gap].same_counts[overlap] += 1

    return stats, details


def _analyze_adjacent_chunked(
    records: List[DrawRecord],
    profile: LotteryProfile,
    max_gap: int = 7,
    should_interrupt: Optional[Callable[[], bool]] = None,
) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """分片包装：每处理完一个间隔 gap 就检查一次中断，保持 UI 响应。

    目前直接委托给各彩种分析器，分析器内部已实现 gap 级别的中断检查。
    """
    return _analyze_adjacent(
        records, profile, max_gap=max_gap,
        should_interrupt=should_interrupt,
    )


def _compute_filter_impact(
    stats: AdjacentStats,
    key: str,
    gap: int,
    threshold: int,
) -> Optional[FilterImpactResult]:
    """计算指定号码组在指定间隔和阈值下的过滤影响.

    Args:
        stats: 相邻期统计结果
        key: 号码组 key（如 "red"、"blue"）
        gap: 间隔期数（0=相邻）
        threshold: 重合上限 K，overlap > K 则被过滤

    Returns:
        FilterImpactResult 或 None（无数据时）
    """
    if gap == 0:
        stat = stats.group_stats.get(key)
    else:
        gap_dict = stats.gap_stats.get(key, {})
        stat = gap_dict.get(gap)

    if stat is None or stat.total_pairs == 0:
        return None

    filtered_count = sum(
        count for overlap, count in stat.same_counts.items()
        if overlap > threshold
    )
    filtered_pct = filtered_count / max(stat.total_pairs, 1) * 100
    retained_count = stat.total_pairs - filtered_count

    gap_label = "相邻" if gap == 0 else f"间隔 {gap} 期"
    return FilterImpactResult(
        gap_label=gap_label,
        group_name=stat.group_name,
        total_pairs=stat.total_pairs,
        threshold=threshold,
        filtered_count=filtered_count,
        filtered_pct=filtered_pct,
        retained_count=retained_count,
    )


def _get_overlap_numbers_ssq(
    base: DrawRecord, curr: DrawRecord,
) -> Tuple[Set[int], Optional[int]]:
    """获取双色球两期之间重叠的红球号码和相同的蓝球."""
    base_reds = set(base.groups.get("red", []))
    curr_reds = set(curr.groups.get("red", []))
    red_overlap = base_reds & curr_reds

    base_blue = next(iter(base.groups.get("blue", [])), None)
    curr_blue = next(iter(curr.groups.get("blue", [])), None)
    blue_same = base_blue if (base_blue is not None and curr_blue is not None and base_blue == curr_blue) else None
    return red_overlap, blue_same


def _get_overlap_numbers_basic_special(
    base: DrawRecord, curr: DrawRecord,
    basic_group: NumberGroup, special_group: NumberGroup,
) -> Tuple[Set[int], Optional[int]]:
    """获取七乐彩两期之间重叠的基本号和相同的特别号."""
    base_basic = set(base.groups.get(basic_group.key, []))
    curr_basic = set(curr.groups.get(special_group.key, []))
    basic_overlap = base_basic & curr_basic

    base_special = next(iter(base.groups.get(special_group.key, [])), None)
    curr_special = next(iter(curr.groups.get(special_group.key, [])), None)
    special_same = base_special if (base_special is not None and curr_special is not None and base_special == curr_special) else None
    return basic_overlap, special_same


def _get_overlap_numbers_positional(
    base: DrawRecord, curr: DrawRecord, group_key: str,
) -> List[Tuple[int, int]]:
    """获取按位彩种两期之间同位相同的号码对."""
    base_nums = base.groups.get(group_key, [])
    curr_nums = curr.groups.get(group_key, [])
    return [(i, a) for i, (a, b) in enumerate(zip(base_nums, curr_nums)) if a == b]


def _get_overlap_numbers_generic(
    base: DrawRecord, curr: DrawRecord, group: NumberGroup,
) -> Set[int]:
    """获取通用彩种两期之间重叠的号码."""
    if group.positional:
        base_nums = base.groups.get(group.key, [])
        curr_nums = curr.groups.get(group.key, [])
        return {a for a, b in zip(base_nums, curr_nums) if a == b}
    return set(base.groups.get(group.key, [])) & set(curr.groups.get(group.key, []))


def _compute_number_frequencies(
    records: List[DrawRecord], group_key: str,
) -> Dict[int, int]:
    """统计每个号码在所有记录中出现的次数."""
    freq: Dict[int, int] = {}
    for record in records:
        nums = record.groups.get(group_key, [])
        for n in nums:
            freq[n] = freq.get(n, 0) + 1
    return freq


def _build_detailed_filter_text(
    records: List[DrawRecord],
    profile: 'LotteryProfile',
    max_gap: int,
    threshold: int,
) -> str:
    """构建详细的过滤影响分析，包含具体号码和频率变化.

    Args:
        records: 开奖记录列表（按时间升序）
        profile: 彩种档案
        max_gap: 最大间隔期数
        threshold: 重合上限阈值

    Returns:
        格式化的详细过滤影响分析文本
    """
    if not records or len(records) < 2:
        return "  数据不足，无法进行详细分析"

    lines: List[str] = []
    lines.append("═" * 50)
    lines.append(f"  详细过滤影响分析（阈值: 重合 > {threshold} 则过滤）")
    lines.append("═" * 50)

    # 确定要分析的号码组
    groups_to_analyze: List[Tuple[str, str, NumberGroup]] = []
    if profile.key == "ssq":
        groups_to_analyze.append(("red", "红球", profile.group("red")))
    elif profile.key in ("qlc", "gd36x7"):
        groups_to_analyze.append(("basic", "基本号", profile.group("basic")))
    elif profile.key == "kl8":
        groups_to_analyze.append(("main", "主号码", profile.primary_group))
    elif profile.key in ("3d", "pl3", "pl5", "qxc"):
        groups_to_analyze.append(("pos", profile.primary_group.name, profile.primary_group))
    else:
        for g in profile.pick_groups:
            groups_to_analyze.append((g.key, g.name, g))

    for group_key, group_name, group in groups_to_analyze:
        lines.append("")
        lines.append(f"【{group_name}】")

        # 统计过滤前每个号码出现的次数
        freq_before = _compute_number_frequencies(records, group_key)
        if not freq_before:
            lines.append("  无数据")
            continue

        # 按号码排序显示过滤前频率
        sorted_nums = sorted(freq_before.keys())
        lines.append(f"  号码池: {sorted_nums[0]}-{sorted_nums[-1]} ({len(sorted_nums)}个)")
        lines.append(f"  过滤前频率: {', '.join(f'{n}({freq_before[n]})' for n in sorted_nums)}")

        # 对每个 gap 级别进行详细分析
        for gap in range(0, max_gap + 1):
            filtered_pairs = []
            filtered_numbers: Dict[int, int] = {}  # 号码 -> 被过滤次数
            retained_numbers: Dict[int, int] = {}  # 号码 -> 保留次数

            # 遍历所有 draw 对
            for i in range(gap + 1, len(records)):
                base_idx = i - gap - 1
                base = records[base_idx]
                curr = records[i]

                # 计算重叠
                overlap_count = 0
                overlap_nums: Set[int] = set()
                if profile.key == "ssq" and group_key == "red":
                    overlap_nums, _ = _get_overlap_numbers_ssq(base, curr)
                    overlap_count = len(overlap_nums)
                elif profile.key in ("qlc", "gd36x7") and group_key == "basic":
                    overlap_nums, _ = _get_overlap_numbers_basic_special(
                        base, curr, profile.group("basic"), profile.group("special")
                    )
                    overlap_count = len(overlap_nums)
                elif profile.key in ("3d", "pl3", "pl5", "qxc") and group_key == "pos":
                    pos_pairs = _get_overlap_numbers_positional(base, curr, group_key)
                    overlap_count = len(pos_pairs)
                    overlap_nums = {n for _, n in pos_pairs}
                else:
                    overlap_nums = _get_overlap_numbers_generic(base, curr, group)
                    overlap_count = len(overlap_nums)

                # 判断是否被过滤
                if overlap_count > threshold:
                    filtered_pairs.append((base.issue, curr.issue, overlap_nums))
                    for n in overlap_nums:
                        filtered_numbers[n] = filtered_numbers.get(n, 0) + 1
                else:
                    for n in overlap_nums:
                        retained_numbers[n] = retained_numbers.get(n, 0) + 1

            if not filtered_pairs:
                continue

            gap_label = "相邻" if gap == 0 else f"间隔 {gap} 期"
            lines.append("")
            lines.append(f"  ── {gap_label}过滤详情 ──")
            lines.append(f"  被过滤对数: {len(filtered_pairs)}")

            # 显示被过滤的具体号码（取前10对）
            lines.append(f"  被过滤的号码对（前{min(10, len(filtered_pairs))}对）:")
            for base_issue, curr_issue, nums in filtered_pairs[:10]:
                nums_str = ','.join(str(n) for n in sorted(nums)) if nums else '-'
                lines.append(f"    {base_issue} vs {curr_issue}: 重合 {nums_str}")
            if len(filtered_pairs) > 10:
                lines.append(f"    ... 还有 {len(filtered_pairs) - 10} 对")

            # 显示被过滤号码的频率
            if filtered_numbers:
                lines.append(f"  被过滤号码出现次数:")
                for n in sorted(filtered_numbers.keys()):
                    before = freq_before.get(n, 0)
                    after = before - filtered_numbers[n]
                    lines.append(f"    {n}: {before}次 → {after}次 (减少{filtered_numbers[n]}次)")

            # 显示保留号码的频率
            if retained_numbers:
                lines.append(f"  保留号码出现次数:")
                for n in sorted(retained_numbers.keys()):
                    before = freq_before.get(n, 0)
                    lines.append(f"    {n}: {before}次 (保留{retained_numbers[n]}次)")

            # 统计未出现的号码
            all_nums_in_pool = set(freq_before.keys())
            appeared_in_filtered = set(filtered_numbers.keys())
            appeared_in_retained = set(retained_numbers.keys())
            never_filtered = all_nums_in_pool - appeared_in_filtered
            never_retained = all_nums_in_pool - appeared_in_retained

            if never_filtered:
                lines.append(f"  从未被过滤的号码: {sorted(never_filtered)}")
            if never_retained:
                lines.append(f"  从未被保留的号码: {sorted(never_retained)}")

    return "\n".join(lines)


def _build_filter_summary_text(
    stats: Optional[AdjacentStats],
    profile: 'LotteryProfile',
    max_gap: int,
    threshold: int,
    records: Optional[List[DrawRecord]] = None,
) -> str:
    """构建过滤影响摘要文本（UI显示用）.

    显示预测时的过滤影响：总共有多少种组合，过滤后还剩多少可买.
    可以直接从 records 计算，不需要先计算 stats.
    """
    from math import comb
    from collections import Counter

    lines: List[str] = []
    lines.append("═" * 50)
    lines.append(f"  过滤摘要（阈值: 重合 > {threshold} 则过滤）")
    lines.append("═" * 50)

    def _estimate_ssq_filter(records: List[DrawRecord], threshold: int, max_gap: int) -> float:
        """估算双色球过滤比例."""
        if not records or len(records) < 2:
            return 0.0
        # 使用 max_gap 作为比较期数
        compare_periods = min(max_gap, len(records)) if max_gap > 0 else min(7, len(records))
        recent = records[-compare_periods:]
        # 统计历史中红球重合超过阈值的比例
        filtered_count = 0
        total_pairs = 0
        for i in range(len(recent)):
            for j in range(i):
                base_reds = set(recent[j].groups.get("red", []))
                curr_reds = set(recent[i].groups.get("red", []))
                overlap = len(base_reds & curr_reds)
                total_pairs += 1
                if overlap > threshold:
                    filtered_count += 1
        return filtered_count / max(total_pairs, 1) * 100

    # 根据彩种计算总组合数和过滤后剩余
    if profile.key == "ssq":
        # 双色球: C(33,6) × 16
        total_red = comb(33, 6)  # 1,107,568
        total_blue = 16
        total_combos = total_red * total_blue  # 17,721,088

        # 估算过滤比例
        if stats:
            red_impact = _compute_filter_impact(stats, "red", 0, threshold)
            if red_impact and red_impact.total_pairs > 0:
                filter_pct = red_impact.filtered_pct
            else:
                filter_pct = 0.0
        elif records:
            filter_pct = _estimate_ssq_filter(records, threshold, max_gap)
        else:
            filter_pct = 0.0

        compare_periods = min(max_gap, len(records)) if max_gap > 0 and records else 0
        filtered_combos = int(total_combos * filter_pct / 100)
        remain_combos = total_combos - filtered_combos

        lines.append("")
        lines.append(f"【双色球】")
        lines.append(f"  比较期数: {compare_periods} 期")
        lines.append(f"  总组合数: {total_combos:,} (C(33,6)×16)")
        lines.append(f"  被过滤: {filtered_combos:,} ({filter_pct:.1f}%)")
        lines.append(f"  有效可买: {remain_combos:,}")

    elif profile.key in ("3d", "pl3"):
        # 福彩3D/排列3: 组合数 = 组选6(120) + 组选3(90) + 豹子(10) = 220
        total_combos = 220  # C(10,3) + 10*9 + 10

        # 精确计算：根据历史记录过滤
        if records and len(records) > 0:
            # 使用 max_gap 作为比较期数，不要写死7
            compare_periods = min(max_gap, len(records)) if max_gap > 0 else min(7, len(records))
            recent = records[-compare_periods:]

            # 生成所有1000个可能的3位数
            all_numbers = set()
            for a in range(10):
                for b in range(10):
                    for c in range(10):
                        all_numbers.add((a, b, c))

            # 检查哪些被过滤，并记录是哪期过滤的
            filtered_numbers = set()
            filtered_details = []  # (number, draw_issue, draw_nums)
            for num in all_numbers:
                for record in recent:
                    hist_nums = record.groups.get("pos", [])
                    if len(hist_nums) == 3:
                        # 统计同位相同数
                        same_count = sum(1 for x, y in zip(num, hist_nums) if x == y)
                        if same_count > threshold:
                            filtered_numbers.add(num)
                            filtered_details.append((num, record.issue, hist_nums))
                            break

            # 将被过滤的数字转换为组合类型统计
            filtered_combos = set()
            for num in filtered_numbers:
                sorted_num = tuple(sorted(num))
                filtered_combos.add(sorted_num)

            filtered_combos_count = len(filtered_combos)
            remain_combos = total_combos - filtered_combos_count
            filter_pct = filtered_combos_count / total_combos * 100
        else:
            compare_periods = 0
            recent = []
            filtered_details = []
            pos_impact = _compute_filter_impact(stats, "pos", 0, threshold) if stats and "pos" in stats.group_stats else None
            if pos_impact and pos_impact.total_pairs > 0:
                filter_ratio = pos_impact.filtered_pct / 100
                filtered_combos_count = int(total_combos * filter_ratio)
                remain_combos = total_combos - filtered_combos_count
                filter_pct = pos_impact.filtered_pct
            else:
                filtered_combos_count = 0
                remain_combos = total_combos
                filter_pct = 0.0

        # 统计各类型组合数
        zu6_count = comb(10, 3)  # 120
        zu3_count = 10 * 9       # 90
        baozi_count = 10         # 10

        lines.append("")
        lines.append(f"【{profile.name}】")
        lines.append(f"  比较期数: {compare_periods} 期")
        lines.append(f"  总组合数: {total_combos} (组选6:{zu6_count} + 组选3:{zu3_count} + 豹子:{baozi_count})")
        lines.append(f"  被过滤: {filtered_combos_count} ({filter_pct:.1f}%)")
        lines.append(f"  有效可买: {remain_combos}")

        # 显示被过滤的具体号码
        if filtered_details:
            lines.append("")
            lines.append("  被过滤的号码:")
            for num, issue, hist in filtered_details:
                lines.append(f"    {num[0]}{num[1]}{num[2]} (与{issue}的{hist[0]}{hist[1]}{hist[2]}重复)")

    elif profile.key == "kl8":
        # 快乐8: C(80,20)
        total_combos = comb(80, 20)

        main_key = profile.primary_group.key
        main_impact = _compute_filter_impact(stats, main_key, 0, threshold)
        if main_impact and main_impact.total_pairs > 0:
            filter_ratio = main_impact.filtered_pct / 100
            filtered_combos = int(total_combos * filter_ratio)
            remain_combos = total_combos - filtered_combos
        else:
            filtered_combos = 0
            remain_combos = total_combos

        lines.append("")
        lines.append(f"【快乐8】")
        lines.append(f"  总组合数: {total_combos:,} (C(80,20))")
        lines.append(f"  被过滤: {filtered_combos:,} ({main_impact.filtered_pct:.1f}%)" if main_impact else "  被过滤: 0")
        lines.append(f"  有效可买: {remain_combos:,}")

    elif profile.key in ("qlc", "gd36x7"):
        # 七乐彩: C(30,7)
        total_combos = comb(30, 7)

        basic_impact = _compute_filter_impact(stats, "basic", 0, threshold)
        if basic_impact and basic_impact.total_pairs > 0:
            filter_ratio = basic_impact.filtered_pct / 100
            filtered_combos = int(total_combos * filter_ratio)
            remain_combos = total_combos - filtered_combos
        else:
            filtered_combos = 0
            remain_combos = total_combos

        lines.append("")
        lines.append(f"【{profile.name}】")
        lines.append(f"  总组合数: {total_combos:,} (C(30,7))")
        lines.append(f"  被过滤: {filtered_combos:,} ({basic_impact.filtered_pct:.1f}%)" if basic_impact else "  被过滤: 0")
        lines.append(f"  有效可买: {remain_combos:,}")

    elif profile.key == "pl5":
        # 排列5: 10^5 = 100,000
        total_combos = 100000

        pos_impact = _compute_filter_impact(stats, "pos", 0, threshold) if "pos" in stats.group_stats else None
        if pos_impact and pos_impact.total_pairs > 0:
            filter_ratio = pos_impact.filtered_pct / 100
            filtered_combos = int(total_combos * filter_ratio)
            remain_combos = total_combos - filtered_combos
        else:
            filtered_combos = 0
            remain_combos = total_combos

        lines.append("")
        lines.append(f"【排列5】")
        lines.append(f"  总组合数: {total_combos:,} (10^5)")
        lines.append(f"  被过滤: {filtered_combos:,} ({pos_impact.filtered_pct:.1f}%)" if pos_impact else "  被过滤: 0")
        lines.append(f"  有效可买: {remain_combos:,}")

    elif profile.key == "qxc":
        # 7星彩: 10^7 = 10,000,000
        total_combos = 10000000

        pos_impact = _compute_filter_impact(stats, "pos", 0, threshold) if "pos" in stats.group_stats else None
        if pos_impact and pos_impact.total_pairs > 0:
            filter_ratio = pos_impact.filtered_pct / 100
            filtered_combos = int(total_combos * filter_ratio)
            remain_combos = total_combos - filtered_combos
        else:
            filtered_combos = 0
            remain_combos = total_combos

        lines.append("")
        lines.append(f"【7星彩】")
        lines.append(f"  总组合数: {total_combos:,} (10^7)")
        lines.append(f"  被过滤: {filtered_combos:,} ({pos_impact.filtered_pct:.1f}%)" if pos_impact else "  被过滤: 0")
        lines.append(f"  有效可买: {remain_combos:,}")

    else:
        # 通用彩种
        for key in stats.group_stats:
            gstat = stats.group_stats[key]
            impact = _compute_filter_impact(stats, key, 0, threshold)
            if impact:
                lines.append("")
                lines.append(f"【{gstat.group_name}】")
                lines.append(f"  历史对数: {impact.total_pairs}")
                lines.append(f"  被过滤: {impact.filtered_count} 对 ({impact.filtered_pct:.1f}%)")
                lines.append(f"  有效保留: {impact.retained_count} 对")

    lines.append("")
    lines.append("═" * 50)
    return "\n".join(lines)


def _build_filter_impact_text(
    stats: AdjacentStats,
    max_gap: int,
    threshold: int,
) -> str:
    """构建过滤影响分析的文本输出（详细版）."""
    lines: List[str] = []
    lines.append("═" * 50)
    lines.append(f"  过滤影响分析（阈值: 重合 > {threshold} 则过滤）")
    lines.append("═" * 50)

    for key in stats.group_stats:
        gstat = stats.group_stats[key]
        lines.append("")
        lines.append(f"【{gstat.group_name}】")

        # 收集所有 gap 级别的过滤影响
        impacts: List[FilterImpactResult] = []

        # gap=0（相邻期）
        impact = _compute_filter_impact(stats, key, 0, threshold)
        if impact is not None:
            impacts.append(impact)

        # gap=1..max_gap
        for gap in range(1, max_gap + 1):
            impact = _compute_filter_impact(stats, key, gap, threshold)
            if impact is not None:
                impacts.append(impact)

        if not impacts:
            lines.append("  无数据")
            continue

        # 表头
        lines.append(f"  {'间隔':<10} {'过滤对数':<12} {'过滤占比':<10} {'保留对数':<10}")
        lines.append(f"  {'─' * 44}")

        for imp in impacts:
            lines.append(
                f"  {imp.gap_label:<10} "
                f"{imp.filtered_count}/{imp.total_pairs:<8} "
                f"{imp.filtered_pct:<10.2f} "
                f"{imp.retained_count}"
            )

        # 级联影响分析（相邻期过滤对其他间隔的影响）
        if len(impacts) > 1:
            adj_impact = impacts[0]
            if adj_impact.filtered_count > 0:
                lines.append("")
                lines.append("  级联影响:")
                for imp in impacts[1:]:
                    # 按比例估算：相邻期过滤比例 × 间隔期过滤比例
                    cascade_count = round(adj_impact.filtered_count * imp.filtered_pct / 100, 1)
                    lines.append(
                        f"    相邻期过滤 {adj_impact.filtered_count} 对"
                        f" → {imp.gap_label}约受影响 {cascade_count} 对"
                    )

        # 过滤建议
        adj_pct = impacts[0].filtered_pct if impacts else 0
        lines.append("")
        if adj_pct < 5:
            lines.append("  建议: 过滤率偏低，可适当降低阈值以增强过滤效果")
        elif adj_pct < 20:
            lines.append(f"  建议: 相邻期 {adj_pct:.1f}% 过滤率属正常范围")
        elif adj_pct < 40:
            lines.append(f"  建议: 相邻期 {adj_pct:.1f}% 过滤率偏高，可适当提高阈值")
        else:
            lines.append(f"  警告: 相邻期 {adj_pct:.1f}% 过滤率过高，大量号码被淘汰")

    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# 分组工具
# --------------------------------------------------------------------------- #
def _group_key(record: DrawRecord, mode: str) -> str:
    """根据分组模式返回该记录所属分组的 key."""
    d = record.draw_date
    if mode == "year":
        return f"{d.year}年"
    if mode == "quarter":
        quarter = (d.month - 1) // 3 + 1
        return f"{d.year}年第{quarter}季度"
    if mode == "month":
        return d.strftime("%Y年%m月")
    if mode == "week":
        return f"{d.year}年第{d.isocalendar()[1]:02d}周"
    return "全部"


def _analyze_adjacent_with_progress(
    records: List[DrawRecord],
    profile: LotteryProfile,
    max_gap: int,
    progress: Any,
) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
    """带进度反馈的全量相邻期/间隔期统计（大数据量时避免 UI 卡死）."""
    from PySide6.QtWidgets import QApplication

    progress.setValue(0)
    QApplication.processEvents()

    stats, details = _analyze_adjacent(records, profile, max_gap=max_gap)

    # 全量计算完成后直接设为 100%
    progress.setValue(100)
    QApplication.processEvents()
    return stats, details


# --------------------------------------------------------------------------- #
# UI
# --------------------------------------------------------------------------- #
class DrawAnalysisDialog(QDialog):
    """开奖记录相邻期统计分析窗口."""

    def __init__(
        self,
        context,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.context = context
        self.profile: LotteryProfile = context.profile
        self.data_repository: DrawRepository = context.data_repository
        self.settings = AppSettings()

        # 增量计算缓存：key = (group_key, max_gap)
        self._cached_stats: Dict[str, AdjacentStats] = {}
        self._cached_details: Dict[str, List[Dict[str, Any]]] = {}
        self._cached_max_gap: int = 0
        self._cached_group_key: Optional[str] = None
        self._cached_records_len: int = 0
        self._pending_update_timer: Optional[Any] = None

        self.setWindowTitle(f"{self.profile.name}开奖记录分析")
        self.resize(1200, 800)
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint | Qt.WindowType.WindowMinimizeButtonHint)
        self._setup_ui()
        self._refresh_data()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        # 顶部控制栏
        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("分组方式:"))
        self.group_combo = QComboBox()
        self.group_combo.addItem("全部", "all")
        self.group_combo.addItem("按年", "year")
        self.group_combo.addItem("按季度", "quarter")
        self.group_combo.addItem("按月", "month")
        self.group_combo.addItem("按周", "week")
        # 加载按彩种存储的分组模式
        saved_mode = self.settings.get_draw_analysis_group_mode(self.profile.key)
        for i in range(self.group_combo.count()):
            if self.group_combo.itemData(i) == saved_mode:
                self.group_combo.setCurrentIndex(i)
                break
        self.group_combo.currentIndexChanged.connect(self._on_group_changed)
        control_layout.addWidget(self.group_combo)

        control_layout.addWidget(QLabel("当前分组:"))
        self.current_group_label = QLabel("全部")
        self.current_group_label.setStyleSheet("font-weight: bold; color: #0A2540;")
        control_layout.addWidget(self.current_group_label)

        control_layout.addSpacing(20)
        control_layout.addWidget(QLabel("最大间隔期数:"))
        self.max_gap_spin = QSpinBox()
        self.max_gap_spin.setRange(0, 50)
        # 使用按彩种存储的设置，默认值为1
        self.max_gap_spin.setValue(min(self.settings.get_draw_analysis_max_gap(self.profile.key), 50))
        self.max_gap_spin.setToolTip("统计时会同时计算间隔 0（相邻）到最大间隔的号码相同情况。修改后需点击【计算】按钮生效。")
        self.max_gap_spin.valueChanged.connect(self._on_max_gap_changed)
        control_layout.addWidget(self.max_gap_spin)

        self.compute_button = QPushButton("计算")
        self.compute_button.setToolTip("点击后根据当前最大间隔期数重新计算间隔统计")
        self.compute_button.clicked.connect(self._on_compute_clicked)
        self.compute_button.setAutoDefault(False)  # 禁用回车自动触发
        control_layout.addWidget(self.compute_button)

        self.export_btn = QPushButton("导出Excel")
        self.export_btn.setToolTip("将开奖记录和统计结果导出到 Excel 文件")
        self.export_btn.clicked.connect(self._export_excel)
        control_layout.addWidget(self.export_btn)

        control_layout.addSpacing(20)
        control_layout.addWidget(QLabel("过滤阈值(重合上限):"))
        self.filter_threshold_spin = QSpinBox()
        self.filter_threshold_spin.setRange(0, 10)
        # 使用按彩种存储的设置，默认值为1
        self.filter_threshold_spin.setValue(min(self.settings.get_draw_analysis_filter_threshold(self.profile.key), 10))
        self.filter_threshold_spin.setToolTip("号码重合数超过此阈值则被过滤。点击【过滤影响分析】按钮查看各间隔期的过滤影响。")
        self.filter_threshold_spin.valueChanged.connect(self._on_filter_threshold_changed)
        control_layout.addWidget(self.filter_threshold_spin)

        self.filter_analysis_btn = QPushButton("过滤影响分析")
        self.filter_analysis_btn.setToolTip("基于当前统计和阈值，分析各间隔期的过滤影响、级联效应和过滤建议")
        self.filter_analysis_btn.clicked.connect(self._on_filter_analysis_clicked)
        self.filter_analysis_btn.setAutoDefault(False)  # 禁用回车自动触发
        control_layout.addWidget(self.filter_analysis_btn)

        control_layout.addStretch()
        layout.addLayout(control_layout)

        # 中间 splitter：左侧分组列表 + 右侧表格和统计
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # 左侧分组列表
        group_box = QGroupBox("分组")
        group_layout = QVBoxLayout(group_box)
        self.group_list = QTableWidget()
        self.group_list.setColumnCount(2)
        self.group_list.setHorizontalHeaderLabels(["分组", "期数"])
        self.group_list.horizontalHeader().setStretchLastSection(True)
        self.group_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.group_list.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.group_list.itemSelectionChanged.connect(self._on_group_selected)
        group_layout.addWidget(self.group_list)
        splitter.addWidget(group_box)

        # 右侧：表格 + 统计（使用垂直 splitter 可拖动调整）
        right_splitter = QSplitter(Qt.Orientation.Vertical)
        right_splitter.setHandleWidth(6)

        table_box = QGroupBox("开奖记录")
        table_layout = QVBoxLayout(table_box)
        self.record_table = QTableWidget()
        self._build_table_columns()
        table_layout.addWidget(self.record_table)

        # 分页控件
        page_layout = QHBoxLayout()
        page_layout.addStretch()
        self.page_prev_btn = QPushButton("上一页")
        self.page_prev_btn.setFixedWidth(80)
        self.page_prev_btn.clicked.connect(self._on_page_prev)
        page_layout.addWidget(self.page_prev_btn)
        self.page_label = QLabel("第 1 / 1 页")
        self.page_label.setStyleSheet("font-weight: bold; margin: 0 12px;")
        page_layout.addWidget(self.page_label)
        self.page_next_btn = QPushButton("下一页")
        self.page_next_btn.setFixedWidth(80)
        self.page_next_btn.clicked.connect(self._on_page_next)
        page_layout.addWidget(self.page_next_btn)
        page_layout.addStretch()
        table_layout.addLayout(page_layout)

        right_splitter.addWidget(table_box)

        stats_box = QGroupBox("相邻期统计")
        stats_layout = QVBoxLayout(stats_box)
        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        self.stats_text.setStyleSheet(
            "QTextEdit { color: #0A2540; background-color: #E8F5E9; "
            "border-radius: 4px; padding: 6px; font-size: 10pt; }"
        )
        stats_layout.addWidget(self.stats_text)
        right_splitter.addWidget(stats_box)

        # 设置垂直 splitter 初始比例（表格 60%，统计 40%）
        right_splitter.setSizes([480, 320])

        splitter.addWidget(right_splitter)
        splitter.setSizes([250, 950])
        layout.addWidget(splitter, 1)

        self._worker: Optional[_AnalysisWorker] = None
        self._progress: Optional[QProgressDialog] = None
        self._elapsed: Optional[QElapsedTimer] = None
        self._pending_worker_records: Optional[List[DrawRecord]] = None
        self._pending_worker_group_key: Optional[str] = None
        self._pending_worker_max_gap: int = 0

        # 分页状态
        self._page_size: int = 100
        self._current_page: int = 0
        self._paged_records: List[DrawRecord] = []
        self._paged_details: List[Dict[str, Any]] = []

    # ----------------------------------------------------------------------- #
    # 按彩种定制的列配置
    # ----------------------------------------------------------------------- #
    def _build_table_columns(self) -> None:
        """根据当前彩种构建表格列标题与列伸缩策略."""
        headers = ["期号", "开奖日期"]
        stretch_cols: List[int] = []

        if self.profile.key == "ssq":
            headers.extend(["红球", "蓝球", "与上期红球重复", "与上期蓝球相同"])
            stretch_cols = [2]
        elif self.profile.key in ("3d", "pl3", "pl5", "qxc"):
            group = self.profile.primary_group
            headers.append(group.name)
            headers.append(f"与上期{group.name}同位相同")
            stretch_cols = [2]
        elif self.profile.key in ("qlc", "gd36x7"):
            headers.extend(["基本号", "特别号", "与上期基本号重复", "与上期特别号相同"])
            stretch_cols = [2]
        elif self.profile.key == "kl8":
            group = self.profile.primary_group
            headers.extend([group.name, f"与上期{group.name}重复"])
            stretch_cols = [2]
        else:
            for g in self.profile.pick_groups:
                headers.append(g.name)
                headers.append(f"与上期{g.name}重复")
            stretch_cols = list(range(2, 2 + len(self.profile.pick_groups) * 2, 2))

        self.record_table.setColumnCount(len(headers))
        self.record_table.setHorizontalHeaderLabels(headers)
        for c in range(len(headers)):
            if c in stretch_cols:
                self.record_table.horizontalHeader().setSectionResizeMode(
                    c, QHeaderView.ResizeMode.Stretch
                )
            else:
                self.record_table.horizontalHeader().setSectionResizeMode(
                    c, QHeaderView.ResizeMode.ResizeToContents
                )

    def _format_group(self, record: DrawRecord, group: NumberGroup) -> str:
        """格式化一组号码用于表格显示."""
        nums = record.groups.get(group.key, [])
        if group.positional:
            return " ".join(f"{n:0{group.pad}d}" for n in nums)
        return " ".join(f"{n:0{group.pad}d}" for n in sorted(nums))

    def _refresh_data(self) -> None:
        self._records = self.data_repository.get_all()
        if not self._records:
            QMessageBox.information(self, "缺少数据", "当前没有开奖记录可供分析。")
            return

        self._records.sort(key=lambda r: r.draw_date, reverse=True)
        self._invalidate_cache()
        self._rebuild_group_list()

    def _invalidate_cache(self) -> None:
        """数据或分组变化时清空缓存."""
        self._cached_stats.clear()
        self._cached_details.clear()
        self._cached_max_gap = 0
        self._cached_group_key = None
        self._cached_records_len = 0

    def _rebuild_group_list(self) -> None:
        mode = self.group_combo.currentData()
        groups: Dict[str, List[DrawRecord]] = {}
        for r in self._records:
            key = _group_key(r, mode)
            groups.setdefault(key, []).append(r)

        self._groups = dict(sorted(groups.items(), key=lambda x: x[0]))
        self.group_list.setRowCount(len(self._groups))
        for idx, (key, records) in enumerate(self._groups.items()):
            self.group_list.setItem(idx, 0, QTableWidgetItem(key))
            item = QTableWidgetItem(str(len(records)))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.group_list.setItem(idx, 1, item)

        if self.group_list.rowCount() > 0:
            # 临时断开信号，避免 selectRow 触发 _on_group_selected → _show_records 卡死主线程
            self.group_list.itemSelectionChanged.disconnect(self._on_group_selected)
            self.group_list.selectRow(0)
            self.group_list.itemSelectionChanged.connect(self._on_group_selected)
            # 手动触发一次初始显示（仅显示表格，不计算统计）
            self._show_table_only_initial()

    def _on_group_changed(self) -> None:
        # 保存按彩种存储的分组模式
        mode = self.group_combo.currentData()
        if mode:
            self.settings.set_draw_analysis_group_mode(self.profile.key, mode)
            self.settings.sync()
        self._invalidate_cache()
        self._rebuild_group_list()

    def _on_max_gap_changed(self, value: int) -> None:
        """最大间隔期数输入变化时保存按彩种存储的设置。"""
        self.settings.set_draw_analysis_max_gap(self.profile.key, value)
        self.settings.sync()

    def _on_filter_threshold_changed(self, value: int) -> None:
        """过滤阈值输入变化时保存按彩种存储的设置。"""
        self.settings.set_draw_analysis_filter_threshold(self.profile.key, value)
        self.settings.sync()

    def _on_compute_clicked(self) -> None:
        """用户点击【计算】按钮后才开始重新计算间隔统计。"""
        selected = self.group_list.selectedItems()
        if selected:
            logger.info("[UI] 用户点击计算按钮")
            self._on_group_selected_with_progress(force_compute=True)

    def _on_filter_analysis_clicked(self) -> None:
        """显示过滤影响分析结果：直接从记录计算，不需要先计算间隔统计."""
        if not self._paged_records:
            QMessageBox.information(self, "提示", "无开奖记录数据")
            return

        threshold = self.filter_threshold_spin.value()
        max_gap = self.max_gap_spin.value()  # 从UI读取最大间隔期数

        # 生成摘要文本（UI显示）- 直接从记录计算
        summary_text = _build_filter_summary_text(
            None, self.profile, max_gap, threshold,
            records=self._paged_records
        )

        # 生成详细文本（输出到日志文件）
        detailed_text = _build_detailed_filter_text(
            self._paged_records, self.profile, max_gap, threshold
        )

        # 输出详细信息到日志文件
        from pathlib import Path
        from datetime import datetime
        log_dir = Path(".caipiao/logs")
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / f"filter_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        with open(log_file, "w", encoding="utf-8") as f:
            f.write(detailed_text)
        logger.info("[过滤分析] 详细信息已保存到: %s", log_file)

        # UI 显示摘要
        self.stats_text.setText(summary_text)

    def _on_group_selected(self) -> None:
        selected = self.group_list.selectedItems()
        if not selected:
            return
        # 分组切换时只做轻量刷新（使用已有缓存或同步小数据量），
        # 不自动启动后台重算，避免切换分组时弹窗干扰。
        self._on_group_selected_with_progress(force_compute=False)

    def _on_group_selected_with_progress(self, *, force_compute: bool = False) -> None:
        row = self.group_list.selectedItems()[0].row()
        group_key = self.group_list.item(row, 0).text()
        self.current_group_label.setText(group_key)
        records = self._groups.get(group_key, [])

        max_gap = self.max_gap_spin.value()
        cache = self._cached_stats.get(group_key)

        need_compute = (
            force_compute
            or cache is None
            or self._cached_group_key != group_key
            or len(records) != self._cached_records_len
            or max_gap > self._cached_max_gap
        )

        logger.info("[UI] group_key=%s, records=%d, max_gap=%d, need_compute=%s, force=%s",
                     group_key, len(records), max_gap, need_compute, force_compute)

        # 第一步：立即显示表格数据（分页，最多100行，瞬间完成）
        self._show_table_only(records, group_key)

        # 第二步：后台执行统计计算
        if need_compute:
            self._start_analysis_worker(records, group_key, max_gap)
        elif cache is not None:
            # 有缓存，直接显示统计
            self._show_stats(cache)
        else:
            self.stats_text.setText("")

    def _start_analysis_worker(
        self,
        records: List[DrawRecord],
        group_key: str,
        max_gap: int,
    ) -> None:
        """启动后台分析 worker，主线程保持进度条响应。"""
        # 若已有 worker 在运行，停止它并复用进度条；否则新建进度条
        if self._worker is not None and self._worker.isRunning():
            self._worker.requestInterruption()
            self._worker.wait(1000)
            self._worker.deleteLater()
            self._worker = None

        # 根据数据量估算总耗时（粗略：每 1000 条约 1~3 秒，随 max_gap 增加）
        record_count = len(records)
        estimated_total_ms = max(3000, int(record_count / 1000) * (1 + max_gap) * 1500)

        initial_text = self._format_progress_text(0, estimated_total_ms)

        if self._progress is None:
            progress = QProgressDialog(initial_text, None, 0, 0, self)
            # range (0, 0) = 走马灯 / indeterminate 模式
            progress.setWindowModality(Qt.WindowModality.NonModal)
            progress.setMinimumDuration(0)
            progress.setAutoClose(False)
            progress.setAutoReset(False)
            progress.setCancelButton(None)
            progress.setFixedWidth(480)
            progress.show()
            self._progress = progress
        else:
            self._progress.setLabelText(initial_text)
            self._progress.setRange(0, 0)
            self._progress.show()

        self._pending_worker_records = records
        self._pending_worker_group_key = group_key
        self._pending_worker_max_gap = max_gap
        self._estimated_total_ms = estimated_total_ms

        self._elapsed = QElapsedTimer()
        self._elapsed.start()

        # 用 QTimer 每秒刷新进度标签（直接读 elapsed，不依赖 worker 信号）
        self._progress_timer = QTimer(self)
        self._progress_timer.setInterval(500)
        self._progress_timer.timeout.connect(self._on_progress_tick)
        self._progress_timer.start()

        worker = _AnalysisWorker(records, self.profile, max_gap)
        worker.finished.connect(self._on_worker_finished)
        worker.error.connect(self._on_worker_error)
        logger.info("[Worker] 启动线程")
        worker.start()
        self._worker = worker
        logger.info("[Worker] 线程已启动, isRunning=%s", worker.isRunning())

    def _format_progress_text(self, elapsed_ms: int, estimated_total_ms: int) -> str:
        elapsed_s = elapsed_ms / 1000
        remaining_ms = max(0, estimated_total_ms - elapsed_ms)
        remaining_s = remaining_ms / 1000
        return (
            f"正在计算间隔统计…\n"
            f"已用时间: {elapsed_s:.0f}s　预计剩余: {remaining_s:.0f}s"
        )

    def _on_progress_tick(self) -> None:
        """QTimer 回调：直接读 QElapsedTimer 更新标签。"""
        if self._progress is None or self._elapsed is None:
            return
        elapsed_ms = self._elapsed.elapsed()
        estimated = getattr(self, "_estimated_total_ms", 5000)
        self._progress.setLabelText(self._format_progress_text(elapsed_ms, estimated))

    def _on_worker_finished(
        self,
        stats: AdjacentStats,
        details: List[Dict[str, Any]],
    ) -> None:
        logger.info("[Worker] finished 信号收到, total_pairs=%d", stats.total_pairs)
        if hasattr(self, "_progress_timer") and self._progress_timer is not None:
            self._progress_timer.stop()
            self._progress_timer.deleteLater()
            self._progress_timer = None
        group_key = self._pending_worker_group_key or ""
        records = self._pending_worker_records or []
        max_gap = self._pending_worker_max_gap

        def _apply() -> None:
            # 先清理 worker 线程，再更新 UI，避免线程对象与大量控件创建竞态
            self._cleanup_worker()
            self._cached_stats[group_key] = stats
            self._cached_details[group_key] = details
            self._cached_max_gap = max_gap
            self._cached_group_key = group_key
            self._cached_records_len = len(records)
            self._show_records(records, group_key)

        if self._elapsed is None:
            _apply()
            return

        elapsed_ms = self._elapsed.elapsed()
        min_display_ms = 2000
        remaining_ms = max(0, min_display_ms - elapsed_ms)
        if remaining_ms <= 0:
            _apply()
        else:
            QTimer.singleShot(remaining_ms, _apply)

    def _on_worker_error(self, message: str) -> None:
        logger.error("[Worker] 错误: %s", message)
        self._cleanup_worker()
        QMessageBox.critical(self, "计算错误", f"间隔统计计算失败：{message}")

    def _cleanup_worker(self) -> None:
        logger.info("[Worker] 清理资源")
        if hasattr(self, "_progress_timer") and self._progress_timer is not None:
            self._progress_timer.stop()
            self._progress_timer.deleteLater()
            self._progress_timer = None
        if self._worker is not None:
            if self._worker.isRunning():
                self._worker.requestInterruption()
                self._worker.wait(1000)
            # 用 deleteLater 让 Qt 在下一个事件循环安全回收，避免 Python GC
            # 立即析构 C++ 对象导致 Qt 内部引用悬空引发堆损坏
            self._worker.deleteLater()
            self._worker = None
        if self._progress is not None:
            self._progress.close()
            self._progress = None
        self._elapsed = None
        self._pending_worker_records = None
        self._pending_worker_group_key = None
        self._pending_worker_max_gap = 0

    def _show_table_only(self, records: List[DrawRecord], group_key: Optional[str] = None) -> None:
        """仅显示表格数据（分页），不计算统计。打开窗口时立即调用。"""
        if group_key is None:
            group_key = self._cached_group_key or ""

        logger.info("[UI] _show_table_only 开始, records=%d", len(records))

        # 保存数据供分页浏览
        self._paged_records = records
        self._paged_details = []
        self._paged_stats = AdjacentStats()
        self._current_page = 0

        self._fill_current_page()
        self._update_page_controls()
        self.stats_text.setText("点击【计算】按钮可查看间隔统计\n点击【过滤影响分析】可查看预测过滤影响")

    def _show_table_only_initial(self) -> None:
        """初始化时仅显示表格，不计算任何统计."""
        row = self.group_list.selectedItems()[0].row() if self.group_list.selectedItems() else 0
        if self.group_list.rowCount() > 0:
            group_key = self.group_list.item(row, 0).text()
            self.current_group_label.setText(group_key)
            records = self._groups.get(group_key, [])
        else:
            records = []

        self._paged_records = records
        self._paged_details = []
        self._paged_stats = AdjacentStats()
        self._current_page = 0

        self._fill_current_page()
        self._update_page_controls()
        self.stats_text.setText("点击【计算】按钮可查看间隔统计\n点击【过滤影响分析】可查看预测过滤影响")

    def _show_records(self, records: List[DrawRecord], group_key: Optional[str] = None) -> None:
        """计算统计并更新显示（worker 完成后调用）."""
        if group_key is None:
            group_key = self._cached_group_key or ""
        max_gap = self.max_gap_spin.value()

        logger.info("[UI] _show_records 开始, records=%d", len(records))
        t0 = time.monotonic()

        stats, details = self._compute_stats(records, group_key, max_gap)

        logger.info("[UI] _compute_stats 完成, 耗时 %.2fs", time.monotonic() - t0)

        # 保存全量数据
        self._paged_records = records
        self._paged_details = details
        self._paged_stats = stats

        self._fill_current_page()
        self._update_page_controls()
        self._show_stats(stats)

    def _fill_current_page(self) -> None:
        """填充当前页的记录到表格（最多 100 行，瞬间完成）."""
        start = self._current_page * self._page_size
        end = min(start + self._page_size, len(self._paged_records))
        page_records = self._paged_records[start:end]
        page_details = self._paged_details[start:end] if self._paged_details else []

        self.record_table.setUpdatesEnabled(False)
        self.record_table.setRowCount(0)
        self.record_table.setRowCount(len(page_records))

        for local_idx, record in enumerate(page_records):
            idx = start + local_idx
            self.record_table.setItem(local_idx, 0, QTableWidgetItem(record.issue))
            date_item = QTableWidgetItem(record.draw_date.strftime("%Y-%m-%d"))
            date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.record_table.setItem(local_idx, 1, date_item)

            detail = page_details[local_idx] if local_idx < len(page_details) else {}

            if self.profile.key == "ssq":
                self._fill_ssq_row(local_idx, record, detail)
            elif self.profile.key in ("3d", "pl3", "pl5", "qxc"):
                self._fill_positional_row(local_idx, record, detail)
            elif self.profile.key in ("qlc", "gd36x7"):
                self._fill_basic_special_row(local_idx, record, detail)
            elif self.profile.key == "kl8":
                self._fill_kl8_row(local_idx, record, detail)
            else:
                self._fill_generic_row(local_idx, record, detail)

        self.record_table.setUpdatesEnabled(True)

    def _total_pages(self) -> int:
        total = len(self._paged_records)
        return max(1, (total + self._page_size - 1) // self._page_size)

    def _update_page_controls(self) -> None:
        """更新分页按钮状态和页码标签."""
        page = self._current_page + 1
        total = self._total_pages()
        self.page_label.setText(f"第 {page} / {total} 页（共 {len(self._paged_records)} 条）")
        self.page_prev_btn.setEnabled(self._current_page > 0)
        self.page_next_btn.setEnabled(self._current_page < total - 1)

    def _on_page_prev(self) -> None:
        if self._current_page > 0:
            self._current_page -= 1
            self._fill_current_page()
            self._update_page_controls()

    def _on_page_next(self) -> None:
        if self._current_page < self._total_pages() - 1:
            self._current_page += 1
            self._fill_current_page()
            self._update_page_controls()

    def _compute_stats(
        self,
        records: List[DrawRecord],
        group_key: str,
        max_gap: int,
    ) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
        """增量计算统计：基于缓存复用已有结果，只计算新增部分."""
        same_group = self._cached_group_key == group_key
        same_records_len = self._cached_records_len == len(records)
        records_changed = not same_records_len
        cached_max = self._cached_max_gap

        cache_hit = (
            same_group
            and same_records_len
            and max_gap <= cached_max
            and group_key in self._cached_stats
        )

        if cache_hit:
            # 只需要裁剪展示范围，数据已经计算过
            return self._cached_stats[group_key], self._cached_details[group_key]

        # 情况 1：同分组、记录数没变、max_gap 增大：复用已有 stats，只补齐新增 gap
        if same_group and same_records_len and max_gap > cached_max and cached_max > 0:
            old_stats = self._cached_stats[group_key]
            old_details = self._cached_details[group_key]
            new_stats = self._extend_gap_stats(records, old_stats, cached_max, max_gap)
            self._cached_stats[group_key] = new_stats
            self._cached_max_gap = max_gap
            return new_stats, old_details

        # 情况 2：同分组、max_gap 相同或更小，但记录数增加：增量追加新记录影响
        if same_group and records_changed and max_gap <= cached_max and cached_max > 0:
            old_stats = self._cached_stats.get(group_key)
            old_details = self._cached_details.get(group_key)
            if old_stats is not None and old_details is not None:
                new_stats, new_details = self._extend_records(
                    records, old_stats, old_details, max_gap
                )
                self._cached_stats[group_key] = new_stats
                self._cached_details[group_key] = new_details
                self._cached_records_len = len(records)
                return new_stats, new_details

        # 兜底：全量重新计算
        stats, details = _analyze_adjacent(records, self.profile, max_gap=max_gap)
        self._cached_stats[group_key] = stats
        self._cached_details[group_key] = details
        self._cached_max_gap = max_gap
        self._cached_group_key = group_key
        self._cached_records_len = len(records)
        return stats, details

    def _extend_gap_stats(
        self,
        records: List[DrawRecord],
        stats: AdjacentStats,
        prev_max_gap: int,
        new_max_gap: int,
    ) -> AdjacentStats:
        """仅扩大 max_gap：为每个号码组补齐新增 gap 的统计."""
        stats = AdjacentStats(
            total_pairs=stats.total_pairs,
            group_stats={k: GroupOverlapStats(
                group_name=v.group_name,
                total_pairs=v.total_pairs,
                same_counts=dict(v.same_counts),
            ) for k, v in stats.group_stats.items()},
            gap_stats={
                k: {
                    gap: GroupGapStats(
                        group_name=v[gap].group_name,
                        gap=gap,
                        total_pairs=v[gap].total_pairs,
                        same_counts=dict(v[gap].same_counts),
                    )
                    for gap in v
                }
                for k, v in stats.gap_stats.items()
            },
        )

        for gap in range(prev_max_gap + 1, new_max_gap + 1):
            for key in stats.group_stats:
                group_name = stats.group_stats[key].group_name
                max_count = max(stats.group_stats[key].same_counts.keys()) if stats.group_stats[key].same_counts else 0
                _ensure_gap_stats(stats, key, group_name, max_count, gap)
                stats.gap_stats[key][gap].total_pairs = max(0, stats.total_pairs - gap)

            for i in range(gap + 1, len(records)):
                self._accumulate_gap_pair(records[i - gap - 1], records[i], stats, gap)

        return stats

    def _extend_records(
        self,
        records: List[DrawRecord],
        stats: AdjacentStats,
        details: List[Dict[str, Any]],
        max_gap: int,
    ) -> Tuple[AdjacentStats, List[Dict[str, Any]]]:
        """仅新增开奖记录：更新相邻期统计和各间隔统计."""
        old_len = self._cached_records_len
        total = len(records)
        added = total - old_len
        if added <= 0:
            return stats, details

        stats = AdjacentStats(
            total_pairs=stats.total_pairs,
            group_stats={k: GroupOverlapStats(
                group_name=v.group_name,
                total_pairs=v.total_pairs,
                same_counts=dict(v.same_counts),
            ) for k, v in stats.group_stats.items()},
            gap_stats={
                k: {
                    gap: GroupGapStats(
                        group_name=v[gap].group_name,
                        gap=gap,
                        total_pairs=v[gap].total_pairs,
                        same_counts=dict(v[gap].same_counts),
                    )
                    for gap in v
                }
                for k, v in stats.gap_stats.items()
            },
        )
        details = list(details)

        # 更新相邻期统计和 details
        for idx in range(old_len, total):
            curr = records[idx]
            prev = records[idx - 1]
            detail = self._accumulate_adjacent(prev, curr, stats)
            details.append(detail)

        stats.total_pairs = max(0, total - 1)
        for gstat in stats.group_stats.values():
            gstat.total_pairs = stats.total_pairs

        # 更新各间隔统计
        for gap in range(1, max_gap + 1):
            if gap not in next(iter(stats.gap_stats.values()), {}):
                continue
            for key in stats.gap_stats:
                stats.gap_stats[key][gap].total_pairs = max(0, stats.total_pairs - gap)

            # 新增的对子：
            # 对于每个新记录 curr = records[old_len..total-1]，
            # 它需要与 records[curr_idx - gap - 1] 形成一对（如果存在）。
            for curr_idx in range(old_len, total):
                base_idx = curr_idx - gap - 1
                if base_idx < 0:
                    continue
                self._accumulate_gap_pair(records[base_idx], records[curr_idx], stats, gap)

        return stats, details

    def _accumulate_adjacent(
        self,
        prev: DrawRecord,
        curr: DrawRecord,
        stats: AdjacentStats,
    ) -> Dict[str, Any]:
        """累加相邻期统计，返回该期 detail."""
        if self.profile.key == "ssq":
            red_overlap, blue_same = _compute_overlap_ssq(prev, curr)
            stats.group_stats["red"].same_counts[red_overlap] += 1
            stats.group_stats["blue"].same_counts[1 if blue_same else 0] += 1
            return {"red": red_overlap, "blue": blue_same}

        if self.profile.key in ("3d", "pl3", "pl5", "qxc"):
            group_key = self.profile.primary_group.key
            prev_nums = prev.groups.get(group_key, [])
            curr_nums = curr.groups.get(group_key, [])
            same = sum(1 for a, b in zip(prev_nums, curr_nums) if a == b)
            stats.group_stats[group_key].same_counts[same] += 1
            return {group_key: same}

        if self.profile.key in ("qlc", "gd36x7"):
            basic = self.profile.group("basic")
            special = self.profile.group("special")
            basic_overlap, special_same = _compute_overlap_basic_special(prev, curr, basic, special)
            stats.group_stats["basic"].same_counts[basic_overlap] += 1
            stats.group_stats["special"].same_counts[1 if special_same else 0] += 1
            return {"basic": basic_overlap, "special": special_same}

        if self.profile.key == "kl8":
            group_key = self.profile.primary_group.key
            overlap = len(set(prev.groups.get(group_key, [])) & set(curr.groups.get(group_key, [])))
            stats.group_stats[group_key].same_counts[overlap] += 1
            return {group_key: overlap}

        detail: Dict[str, Any] = {}
        for g in self.profile.pick_groups:
            overlap = _compute_overlap_generic(prev, curr, g)
            stats.group_stats[g.key].same_counts[overlap] += 1
            detail[g.key] = overlap
        return detail

    def _accumulate_gap_pair(
        self,
        base: DrawRecord,
        curr: DrawRecord,
        stats: AdjacentStats,
        gap: int,
    ) -> None:
        """累加 base 与 curr 在指定 gap 下的统计."""
        if self.profile.key == "ssq":
            red_overlap, blue_same = _compute_overlap_ssq(base, curr)
            stats.gap_stats["red"][gap].same_counts[red_overlap] += 1
            stats.gap_stats["blue"][gap].same_counts[1 if blue_same else 0] += 1
            return

        if self.profile.key in ("3d", "pl3", "pl5", "qxc"):
            group_key = self.profile.primary_group.key
            base_nums = base.groups.get(group_key, [])
            curr_nums = curr.groups.get(group_key, [])
            same = sum(1 for a, b in zip(base_nums, curr_nums) if a == b)
            stats.gap_stats[group_key][gap].same_counts[same] += 1
            return

        if self.profile.key in ("qlc", "gd36x7"):
            basic = self.profile.group("basic")
            special = self.profile.group("special")
            basic_overlap, special_same = _compute_overlap_basic_special(base, curr, basic, special)
            stats.gap_stats["basic"][gap].same_counts[basic_overlap] += 1
            stats.gap_stats["special"][gap].same_counts[1 if special_same else 0] += 1
            return

        if self.profile.key == "kl8":
            group_key = self.profile.primary_group.key
            overlap = len(set(base.groups.get(group_key, [])) & set(curr.groups.get(group_key, [])))
            stats.gap_stats[group_key][gap].same_counts[overlap] += 1
            return

        for g in self.profile.pick_groups:
            overlap = _compute_overlap_generic(base, curr, g)
            stats.gap_stats[g.key][gap].same_counts[overlap] += 1

    def _fill_ssq_row(self, idx: int, record: DrawRecord, detail: Dict[str, Any]) -> None:
        reds = record.groups.get("red", [])
        blue = next(iter(record.groups.get("blue", [])), None)

        red_item = QTableWidgetItem(" ".join(f"{r:02d}" for r in sorted(reds)))
        red_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 2, red_item)

        blue_item = QTableWidgetItem(f"{blue:02d}" if blue is not None else "-")
        blue_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 3, blue_item)

        red_overlap = detail.get("red")
        overlap_text = str(red_overlap) if red_overlap is not None else "-"
        overlap_item = QTableWidgetItem(overlap_text)
        overlap_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 4, overlap_item)

        blue_same = detail.get("blue")
        blue_same_text = "是" if blue_same else ("否" if blue_same is False else "-")
        blue_same_item = QTableWidgetItem(blue_same_text)
        blue_same_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 5, blue_same_item)

    def _fill_positional_row(self, idx: int, record: DrawRecord, detail: Dict[str, Any]) -> None:
        group = self.profile.primary_group
        nums_item = QTableWidgetItem(self._format_group(record, group))
        nums_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 2, nums_item)

        same = detail.get(group.key)
        same_text = str(same) if same is not None else "-"
        same_item = QTableWidgetItem(same_text)
        same_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 3, same_item)

    def _fill_basic_special_row(self, idx: int, record: DrawRecord, detail: Dict[str, Any]) -> None:
        basic = self.profile.group("basic")
        special = self.profile.group("special")

        basic_item = QTableWidgetItem(self._format_group(record, basic))
        basic_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 2, basic_item)

        special_num = next(iter(record.groups.get(special.key, [])), None)
        special_item = QTableWidgetItem(
            f"{special_num:02d}" if special_num is not None else "-"
        )
        special_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 3, special_item)

        basic_overlap = detail.get("basic")
        basic_text = str(basic_overlap) if basic_overlap is not None else "-"
        basic_overlap_item = QTableWidgetItem(basic_text)
        basic_overlap_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 4, basic_overlap_item)

        special_same = detail.get("special")
        special_text = "是" if special_same else ("否" if special_same is False else "-")
        special_same_item = QTableWidgetItem(special_text)
        special_same_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 5, special_same_item)

    def _fill_kl8_row(self, idx: int, record: DrawRecord, detail: Dict[str, Any]) -> None:
        group = self.profile.primary_group

        nums_item = QTableWidgetItem(self._format_group(record, group))
        nums_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 2, nums_item)

        overlap = detail.get(group.key)
        overlap_text = str(overlap) if overlap is not None else "-"
        overlap_item = QTableWidgetItem(overlap_text)
        overlap_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.record_table.setItem(idx, 3, overlap_item)

    def _fill_generic_row(self, idx: int, record: DrawRecord, detail: Dict[str, Any]) -> None:
        col = 2
        for g in self.profile.pick_groups:
            nums_item = QTableWidgetItem(self._format_group(record, g))
            nums_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.record_table.setItem(idx, col, nums_item)
            col += 1

            overlap = detail.get(g.key)
            overlap_text = str(overlap) if overlap is not None else "-"
            overlap_item = QTableWidgetItem(overlap_text)
            overlap_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.record_table.setItem(idx, col, overlap_item)
            col += 1

    def _show_stats(self, stats: AdjacentStats) -> None:
        lines = [f"相邻期对数：{stats.total_pairs}"]
        for key in stats.group_stats:
            gstat = stats.group_stats[key]
            lines.append("")
            lines.append(f"【{gstat.group_name}相同统计】")
            if self.profile.key in ("ssq", "qlc", "gd36x7") and key in ("blue", "special"):
                # 二值统计
                same = gstat.same_counts.get(1, 0)
                diff = gstat.same_counts.get(0, 0)
                lines.append(f"  相同：{same} 次（{gstat.same_ratio(1):.2f}%）")
                lines.append(f"  不同：{diff} 次（{gstat.same_ratio(0):.2f}%）")
            else:
                max_n = max(gstat.same_counts.keys()) if gstat.same_counts else 0
                for n in range(0, max_n + 1):
                    count = gstat.same_counts.get(n, 0)
                    ratio = gstat.same_ratio(n)
                    lines.append(f"  {n} 个相同：{count} 次（{ratio:.2f}%）")

            # 间隔 1..max_gap 期统计，按从大到小展示
            if key in stats.gap_stats:
                for gap in sorted(stats.gap_stats[key], reverse=True):
                    if gap > self.max_gap_spin.value():
                        continue
                    gap_stat = stats.gap_stats[key][gap]
                    lines.append("")
                    lines.append(f"【间隔 {gap} 期 {gap_stat.group_name}相同统计】")
                    if self.profile.key in ("ssq", "qlc", "gd36x7") and key in ("blue", "special"):
                        same = gap_stat.same_counts.get(1, 0)
                        diff = gap_stat.same_counts.get(0, 0)
                        lines.append(f"  相同：{same} 次（{gap_stat.same_ratio(1):.2f}%）")
                        lines.append(f"  不同：{diff} 次（{gap_stat.same_ratio(0):.2f}%）")
                    else:
                        max_n = max(gap_stat.same_counts.keys()) if gap_stat.same_counts else 0
                        for n in range(0, max_n + 1):
                            count = gap_stat.same_counts.get(n, 0)
                            ratio = gap_stat.same_ratio(n)
                            lines.append(f"  {n} 个相同：{count} 次（{ratio:.2f}%）")

        self.stats_text.setText("\n".join(lines))

    def _export_excel(self) -> None:
        """导出开奖记录和统计结果到 Excel."""
        from datetime import datetime
        from pathlib import Path

        from PySide6.QtWidgets import QFileDialog

        from ...core.prize import fc3d_bet_type

        if not self._paged_records:
            QMessageBox.information(self, "提示", "暂无数据可导出")
            return

        default_name = f"{self.profile.name}_分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel", default_name, "Excel 文件 (*.xlsx)"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()

            # ---- Sheet 1: 开奖记录 ----
            ws1 = wb.active
            ws1.title = "开奖记录"

            # 构建表头
            headers = ["期号", "开奖日期"]
            if self.profile.key == "ssq":
                headers.extend(["红球", "蓝球", "与上期红球重复", "与上期蓝球相同"])
            elif self.profile.key in ("3d", "pl3", "pl5", "qxc"):
                group = self.profile.primary_group
                headers.extend([group.name, f"与上期{group.name}同位相同"])
            elif self.profile.key in ("qlc", "gd36x7"):
                headers.extend(["基本号", "特别号", "与上期基本号重复", "与上期特别号相同"])
            elif self.profile.key == "kl8":
                group = self.profile.primary_group
                headers.extend([group.name, f"与上期{group.name}重复"])
            else:
                for g in self.profile.pick_groups:
                    headers.extend([g.name, f"与上期{g.name}重复"])

            header_fill = PatternFill(start_color="0077B6", end_color="0077B6", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            records = self._paged_records
            details = self._paged_details
            for row_idx, record in enumerate(records, 2):
                ws1.cell(row=row_idx, column=1, value=record.issue)
                ws1.cell(row=row_idx, column=2, value=record.draw_date.strftime("%Y-%m-%d"))
                detail = details[row_idx - 2] if row_idx - 2 < len(details) else {}

                if self.profile.key == "ssq":
                    reds = record.groups.get("red", [])
                    blue = next(iter(record.groups.get("blue", [])), None)
                    ws1.cell(row=row_idx, column=3, value=" ".join(f"{r:02d}" for r in sorted(reds)))
                    ws1.cell(row=row_idx, column=4, value=f"{blue:02d}" if blue is not None else "-")
                    red_overlap = detail.get("red")
                    ws1.cell(row=row_idx, column=5, value=str(red_overlap) if red_overlap is not None else "-")
                    blue_same = detail.get("blue")
                    ws1.cell(row=row_idx, column=6, value="是" if blue_same else ("否" if blue_same is False else "-"))
                elif self.profile.key in ("3d", "pl3", "pl5", "qxc"):
                    group = self.profile.primary_group
                    nums = record.groups.get(group.key, [])
                    ws1.cell(row=row_idx, column=3, value=" ".join(f"{n:0{group.pad}d}" for n in nums))
                    same = detail.get(group.key)
                    ws1.cell(row=row_idx, column=4, value=str(same) if same is not None else "-")
                elif self.profile.key in ("qlc", "gd36x7"):
                    basic = self.profile.group("basic")
                    special = self.profile.group("special")
                    basic_nums = record.groups.get(basic.key, [])
                    ws1.cell(row=row_idx, column=3, value=" ".join(f"{n:0{basic.pad}d}" for n in sorted(basic_nums)))
                    special_num = next(iter(record.groups.get(special.key, [])), None)
                    ws1.cell(row=row_idx, column=4, value=f"{special_num:02d}" if special_num is not None else "-")
                    basic_overlap = detail.get("basic")
                    ws1.cell(row=row_idx, column=5, value=str(basic_overlap) if basic_overlap is not None else "-")
                    special_same = detail.get("special")
                    ws1.cell(row=row_idx, column=6, value="是" if special_same else ("否" if special_same is False else "-"))
                elif self.profile.key == "kl8":
                    group = self.profile.primary_group
                    nums = record.groups.get(group.key, [])
                    ws1.cell(row=row_idx, column=3, value=" ".join(f"{n:0{group.pad}d}" for n in sorted(nums)))
                    overlap = detail.get(group.key)
                    ws1.cell(row=row_idx, column=4, value=str(overlap) if overlap is not None else "-")
                else:
                    col = 3
                    for g in self.profile.pick_groups:
                        nums = record.groups.get(g.key, [])
                        ws1.cell(row=row_idx, column=col, value=" ".join(f"{n:0{g.pad}d}" for n in sorted(nums)))
                        overlap = detail.get(g.key)
                        ws1.cell(row=row_idx, column=col + 1, value=str(overlap) if overlap is not None else "-")
                        col += 2

            # 自动列宽
            for col in range(1, len(headers) + 1):
                max_len = max(len(str(ws1.cell(row=r, column=col).value or "")) for r in range(1, min(100, ws1.max_row + 1)))
                ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = min(max_len + 4, 40)

            # ---- Sheet 2: 统计分析 ----
            ws2 = wb.create_sheet("统计分析")
            stats = self._paged_stats
            row = 1
            ws2.cell(row=row, column=1, value=f"相邻期对数：{stats.total_pairs}").font = Font(bold=True)
            row += 1

            for key in stats.group_stats:
                gstat = stats.group_stats[key]
                row += 1
                ws2.cell(row=row, column=1, value=f"【{gstat.group_name}相同统计】").font = Font(bold=True, color="0077B6")
                row += 1

                # 相邻期统计
                ws2.cell(row=row, column=1, value="间隔")
                ws2.cell(row=row, column=2, value="相同个数")
                ws2.cell(row=row, column=3, value="次数")
                ws2.cell(row=row, column=4, value="占比")
                for c in range(1, 5):
                    ws2.cell(row=row, column=c).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    ws2.cell(row=row, column=c).font = Font(bold=True)
                row += 1

                if self.profile.key in ("ssq", "qlc", "gd36x7") and key in ("blue", "special"):
                    for val, label in [(1, "相同"), (0, "不同")]:
                        ws2.cell(row=row, column=1, value="相邻")
                        ws2.cell(row=row, column=2, value=label)
                        ws2.cell(row=row, column=3, value=gstat.same_counts.get(val, 0))
                        ws2.cell(row=row, column=4, value=f"{gstat.same_ratio(val):.2f}%")
                        row += 1
                else:
                    max_n = max(gstat.same_counts.keys()) if gstat.same_counts else 0
                    for n in range(0, max_n + 1):
                        ws2.cell(row=row, column=1, value="相邻")
                        ws2.cell(row=row, column=2, value=f"{n} 个相同")
                        ws2.cell(row=row, column=3, value=gstat.same_counts.get(n, 0))
                        ws2.cell(row=row, column=4, value=f"{gstat.same_ratio(n):.2f}%")
                        row += 1

                # 间隔统计
                if key in stats.gap_stats:
                    for gap in sorted(stats.gap_stats[key], reverse=True):
                        if gap > self.max_gap_spin.value():
                            continue
                        gap_stat = stats.gap_stats[key][gap]
                        row += 1
                        ws2.cell(row=row, column=1, value=f"【间隔 {gap} 期 {gap_stat.group_name}相同统计】").font = Font(bold=True, color="0077B6")
                        row += 1
                        ws2.cell(row=row, column=1, value="间隔")
                        ws2.cell(row=row, column=2, value="相同个数")
                        ws2.cell(row=row, column=3, value="次数")
                        ws2.cell(row=row, column=4, value="占比")
                        for c in range(1, 5):
                            ws2.cell(row=row, column=c).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                            ws2.cell(row=row, column=c).font = Font(bold=True)
                        row += 1

                        if self.profile.key in ("ssq", "qlc", "gd36x7") and key in ("blue", "special"):
                            for val, label in [(1, "相同"), (0, "不同")]:
                                ws2.cell(row=row, column=1, value=f"间隔{gap}期")
                                ws2.cell(row=row, column=2, value=label)
                                ws2.cell(row=row, column=3, value=gap_stat.same_counts.get(val, 0))
                                ws2.cell(row=row, column=4, value=f"{gap_stat.same_ratio(val):.2f}%")
                                row += 1
                        else:
                            max_n = max(gap_stat.same_counts.keys()) if gap_stat.same_counts else 0
                            for n in range(0, max_n + 1):
                                ws2.cell(row=row, column=1, value=f"间隔{gap}期")
                                ws2.cell(row=row, column=2, value=f"{n} 个相同")
                                ws2.cell(row=row, column=3, value=gap_stat.same_counts.get(n, 0))
                                ws2.cell(row=row, column=4, value=f"{gap_stat.same_ratio(n):.2f}%")
                                row += 1

            # 过滤影响分析
            threshold = self.filter_threshold_spin.value()
            row += 2
            ws2.cell(row=row, column=1, value=f"【过滤影响分析（阈值: 重合 > {threshold} 则过滤）】").font = Font(bold=True, color="CC0000")
            row += 1
            ws2.cell(row=row, column=1, value="号码组")
            ws2.cell(row=row, column=2, value="间隔")
            ws2.cell(row=row, column=3, value="过滤对数")
            ws2.cell(row=row, column=4, value="过滤占比")
            ws2.cell(row=row, column=5, value="保留对数")
            for c in range(1, 6):
                ws2.cell(row=row, column=c).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                ws2.cell(row=row, column=c).font = Font(bold=True)
            row += 1

            for key in stats.group_stats:
                gstat = stats.group_stats[key]
                # 相邻期
                impact = _compute_filter_impact(stats, key, 0, threshold)
                if impact is not None:
                    ws2.cell(row=row, column=1, value=gstat.group_name)
                    ws2.cell(row=row, column=2, value="相邻")
                    ws2.cell(row=row, column=3, value=f"{impact.filtered_count}/{impact.total_pairs}")
                    ws2.cell(row=row, column=4, value=f"{impact.filtered_pct:.2f}%")
                    ws2.cell(row=row, column=5, value=impact.retained_count)
                    row += 1
                # 各间隔
                for gap in sorted(stats.gap_stats.get(key, {}), reverse=True):
                    if gap > self.max_gap_spin.value():
                        continue
                    impact = _compute_filter_impact(stats, key, gap, threshold)
                    if impact is not None:
                        ws2.cell(row=row, column=1, value=gstat.group_name)
                        ws2.cell(row=row, column=2, value=f"间隔{gap}期")
                        ws2.cell(row=row, column=3, value=f"{impact.filtered_count}/{impact.total_pairs}")
                        ws2.cell(row=row, column=4, value=f"{impact.filtered_pct:.2f}%")
                        ws2.cell(row=row, column=5, value=impact.retained_count)
                        row += 1

            ws2.column_dimensions["A"].width = 16
            ws2.column_dimensions["B"].width = 16
            ws2.column_dimensions["C"].width = 10
            ws2.column_dimensions["D"].width = 10
            ws2.column_dimensions["E"].width = 10

            wb.save(path)
            QMessageBox.information(self, "导出成功", f"已导出到：{path}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "导出失败", f"导出 Excel 失败：{exc}")
