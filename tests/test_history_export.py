"""HistoryManager 导出功能测试."""

from datetime import datetime
from pathlib import Path

import pytest

from caipiao.core.ticket import Ticket
from caipiao.persistence.history import HistoryManager


# ---- Helpers ----

def _make_tickets(count=5):
    tickets = []
    for i in range(count):
        tickets.append(Ticket(
            red_balls=[i + 1, i + 2, i + 3, i + 4, i + 5, i + 6],
            blue_ball=(i % 16) + 1,
            generated_at=datetime(2024, 1, 1, 12, 0, i),
            strategy_name="test_strategy",
            basis=f"test_basis_{i}",
        ))
    return tickets


# ---- CSV Export Tests ----

class TestExportCSV:
    """CSV 导出测试."""

    def test_export_csv(self, tmp_path):
        path = tmp_path / "history.csv"
        hm = HistoryManager(tmp_path / "history.json")
        hm.add_many(_make_tickets(5))
        hm.export_csv(path)
        assert path.exists()
        content = path.read_text(encoding="utf-8-sig")
        assert "时间" in content
        assert "策略" in content
        assert "test_strategy" in content

    def test_export_csv_empty(self, tmp_path):
        path = tmp_path / "history.csv"
        hm = HistoryManager(tmp_path / "history.json")
        hm.export_csv(path)
        assert path.exists()
        content = path.read_text(encoding="utf-8-sig")
        assert "时间" in content


# ---- TXT Export Tests ----

class TestExportTXT:
    """TXT 导出测试."""

    def test_export_txt(self, tmp_path):
        path = tmp_path / "history.txt"
        hm = HistoryManager(tmp_path / "history.json")
        hm.add_many(_make_tickets(5))
        hm.export_txt(path)
        assert path.exists()
        content = path.read_text(encoding="utf-8")
        lines = content.strip().split("\n")
        assert len(lines) == 5

    def test_export_txt_empty(self, tmp_path):
        path = tmp_path / "history.txt"
        hm = HistoryManager(tmp_path / "history.json")
        hm.export_txt(path)
        assert path.exists()
        assert path.read_text(encoding="utf-8") == ""


# ---- Excel Export Tests ----

class TestExportExcel:
    """Excel 导出测试."""

    def test_export_excel(self, tmp_path):
        pytest.importorskip("openpyxl")
        path = tmp_path / "history.xlsx"
        hm = HistoryManager(tmp_path / "history.json")
        hm.add_many(_make_tickets(5))
        hm.export_excel(path)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_export_excel_content(self, tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        path = tmp_path / "history.xlsx"
        hm = HistoryManager(tmp_path / "history.json")
        hm.add_many(_make_tickets(3))
        hm.export_excel(path)

        wb = load_workbook(path)
        ws = wb.active
        # 表头 + 3 行数据
        assert ws.max_row == 4
        # 检查表头
        assert ws.cell(row=1, column=1).value == "序号"
        assert ws.cell(row=1, column=2).value == "时间"
        assert ws.cell(row=1, column=3).value == "策略"
        # 检查数据
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=3).value == "test_strategy"

    def test_export_excel_empty(self, tmp_path):
        pytest.importorskip("openpyxl")
        path = tmp_path / "history.xlsx"
        hm = HistoryManager(tmp_path / "history.json")
        hm.export_excel(path)
        assert path.exists()

    def test_export_excel_freeze_panes(self, tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        path = tmp_path / "history.xlsx"
        hm = HistoryManager(tmp_path / "history.json")
        hm.add_many(_make_tickets(1))
        hm.export_excel(path)

        wb = load_workbook(path)
        ws = wb.active
        assert ws.freeze_panes == "A2"


# ---- Import Tests ----

class TestImportJSON:
    """JSON 导入测试."""

    def test_import_json(self, tmp_path):
        import json

        # 创建导出文件
        source = tmp_path / "source.json"
        tickets = _make_tickets(3)
        data = [t.to_dict() for t in tickets]
        source.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        # 导入
        hm = HistoryManager(tmp_path / "history.json")
        count = hm.import_from_json(source)
        assert count == 3
        assert hm.get_all().__len__() == 3

    def test_import_json_invalid(self, tmp_path):
        source = tmp_path / "invalid.json"
        source.write_text("not json", encoding="utf-8")
        hm = HistoryManager(tmp_path / "history.json")
        count = hm.import_from_json(source)
        assert count == 0
